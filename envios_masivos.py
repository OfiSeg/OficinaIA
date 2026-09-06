from __future__ import annotations

import csv
import io
import json
import os
import re
import shutil
import tempfile
import time
import unicodedata
import uuid
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape as _xml_escape

from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parent
PLANTILLA_ENVIOSYA = BASE_DIR / "plantillas" / "enviosya_contactos.xlsx"
TMP_DIR = Path(tempfile.gettempdir()) / "oficinaia_envios_masivos"
TMP_DIR.mkdir(parents=True, exist_ok=True)
TTL_TEMP_SEGUNDOS = 60 * 60
MAX_ARCHIVOS = 12
MAX_FILAS_POR_ARCHIVO = 120_000

CAMPOS_SALIDA = [
    "apellido", "nombre", "dni", "celular", "localidad", "compania",
    "patente", "marca", "modelo", "anio", "cliente", "vencimiento",
]

ALIASES_CAMPOS = {
    "apellido": {"apellido", "apellidos", "surname", "last name"},
    "nombre": {"nombre", "nombres", "first name", "name"},
    "nombre_completo": {"nombre completo", "apellido y nombre", "apellidos y nombres", "cliente nombre", "razon social", "razón social", "titular"},
    "dni": {"dni", "documento", "nro doc", "nro. doc", "nro documento", "numero documento", "número documento", "doc"},
    "celular": {"celular", "telefono", "teléfono", "telefono celular", "tel celular", "tel", "movil", "móvil", "whatsapp", "cel", "nro celular", "numero celular", "número celular"},
    "localidad": {"localidad", "ciudad", "city", "poblacion", "población"},
    "cp": {"cp", "codigo postal", "código postal", "cod postal", "postal", "cpa"},
    "direccion": {"direccion", "dirección", "domicilio", "calle", "address"},
    "compania": {"compania", "compañia", "compañía", "aseguradora", "cia", "cia seguro", "compania seguro"},
    "patente": {"patente", "dominio", "chapa", "registration", "matricula", "matrícula"},
    "marca": {"marca", "brand"},
    "modelo": {"modelo", "model"},
    "anio": {"año", "anio", "year", "modelo año", "modelo anio"},
    "cliente": {"cliente", "productor", "organizador", "broker"},
    "vencimiento": {"vencimiento", "fecha vencimiento", "vigencia hasta", "hasta", "renovacion", "renovación"},
    "poliza": {"poliza", "póliza", "nro poliza", "nro. poliza", "numero poliza", "número póliza", "policy", "policy number"},
    "provincia": {"provincia", "province", "estado"},
    "pais": {"pais", "país", "country"},
    "fecha_origen": {"fecha", "fecha origen", "fecha alta", "alta"},
    "email": {"email", "e-mail", "mail", "correo", "correo electronico", "correo electrónico"},
    "tipo": {"tipo", "ramo", "producto", "categoria", "categoría", "segmento"},
}

TIPOS_CONOCIDOS = {
    "VEHICULAR", "AUTOMOTOR", "AUTO", "MOTO", "COMERCIO", "DOMICILIO",
    "HOGAR", "FLOTA", "OBJETO", "RC", "RESPONSABILIDAD CIVIL", "AP",
    "ACCIDENTES PERSONALES", "CONSORCIO", "BICICLETA",
}

NOMBRES_COMUNES = {
    "ABEL", "ABRAHAM", "ADRIAN", "ADRIANA", "AGUSTIN", "AGUSTINA", "ALAN", "ALBERTO", "ALEJANDRA", "ALEJANDRO",
    "ALEXIS", "ALICIA", "ALMA", "AMALIA", "AMANDA", "AMELIA", "ANA", "ANDREA", "ANDRES", "ANGEL", "ANGELA", "ANTONELA",
    "ANTONIA", "ANTONIO", "ARIEL", "ARMANDO", "BEATRIZ", "BELEN", "BENJAMIN", "BERNARDO", "BRENDA", "BRIAN", "BRUNO",
    "CAMILA", "CARINA", "CARLA", "CARLOS", "CARMEN", "CAROLINA", "CATALINA", "CECILIA", "CESAR", "CLAUDIA", "CLAUDIO",
    "CRISTIAN", "CRISTINA", "DANIEL", "DANIELA", "DARIO", "DAVID", "DEBORA", "DIEGO", "EDGARDO", "EDITH", "EDUARDO",
    "ELENA", "ELIAS", "EMILIA", "EMILIANO", "EMILIO", "ENRIQUE", "ESTEBAN", "EUGENIA", "EZEQUIEL", "FABIANA", "FABIAN",
    "FABIO", "FACUNDO", "FEDERICO", "FELIPE", "FERNANDA", "FERNANDO", "FLORENCIA", "FRANCO", "GABRIEL", "GABRIELA",
    "GERMAN", "GLADIS", "GLORIA", "GONZALO", "GRACIELA", "GRISELDA", "GUADALUPE", "GUILLERMO", "GUSTAVO", "HECTOR",
    "HERNAN", "HILDA", "HORACIO", "HUGO", "INES", "ISABEL", "IVAN", "JAVIER", "JESICA", "JOAQUIN", "JORGE", "JOSE",
    "JOSEFINA", "JUAN", "JUANA", "JULIA", "JULIANA", "JULIO", "KAREN", "KARINA", "LAURA", "LEANDRO", "LEONARDO",
    "LETICIA", "LILIANA", "LORENA", "LOURDES", "LUCAS", "LUCIANA", "LUCIANO", "LUIS", "LUISA", "LUZ", "MAGALI",
    "MANUEL", "MARCELA", "MARCELO", "MARCOS", "MARGARITA", "MARIA", "MARIANA", "MARIANO", "MARIELA", "MARINA", "MARIO",
    "MARTA", "MARTIN", "MATIAS", "MAURICIO", "MAXIMILIANO", "MELANIE", "MERCEDES", "MICAELA", "MIGUEL", "MILAGROS",
    "MIRTA", "MONICA", "NAHUEL", "NATALIA", "NESTOR", "NICOLAS", "NOELIA", "NORMA", "OSCAR", "PABLO", "PAMELA",
    "PATRICIA", "PAULA", "PEDRO", "RAFAEL", "RAMIRO", "RAMON", "RAUL", "REBECA", "RICARDO", "ROBERTO", "ROCIO",
    "RODOLFO", "RODRIGO", "ROMINA", "ROSA", "ROSANA", "RUBEN", "SABRINA", "SANDRA", "SANTIAGO", "SARA", "SEBASTIAN",
    "SERGIO", "SILVANA", "SILVIA", "SOFIA", "SOL", "SONIA", "STELLA", "SUSANA", "TAMARA", "TATIANA", "VALENTINA",
    "VALERIA", "VERONICA", "VICTOR", "VICTORIA", "VIVIANA", "WALTER", "YANINA",
}

COMPANIAS_FILENAME = {
    "ALLIANZ": "ALLIANZ",
    "ATM": "ATM SEGUROS",
    "AGROSALTA": "AGROSALTA SEGUROS",
    "FEDERACION PATRONAL": "FEDERACION PATRONAL",
    "FEDERACIÓN PATRONAL": "FEDERACION PATRONAL",
    "MERCANTIL ANDINA": "MERCANTIL ANDINA",
    "SAN CRISTOBAL": "SAN CRISTOBAL",
    "SAN CRISTÓBAL": "SAN CRISTOBAL",
    "RIVADAVIA": "RIVADAVIA",
    "TRIUNFO": "TRIUNFO SEGUROS",
    "PROF": "PROF SEGUROS",
    "EUROAMERICA": "EUROAMERICA",
    "EUROAMÉRICA": "EUROAMERICA",
}


def _norm_texto(v: Any) -> str:
    s = str(v or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s.upper().strip()


def _texto_limpio(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return re.sub(r"\s+", " ", str(v).strip())


def _reparar_texto_fuente(v: Any, campo: str = "") -> tuple[str, bool, bool]:
    """Limpia daños conocidos de bases heredadas sin tocar datos sensibles.

    En la muestra Allianz el propio XLSX de origen trae '#' en lugar de Ñ
    (MU#OZ, CA#UELAS, etc.). Sólo se repara en campos de texto humano donde
    ese carácter no tiene uso normal. Si queda U+FFFD (�), se marca para
    revisión en vez de inventar el carácter original.
    """
    texto = _texto_limpio(v)
    reparado = False
    campos_naturales = {
        "apellido", "nombre", "nombre_completo", "localidad", "direccion",
        "provincia", "cliente", "marca", "modelo", "tipo",
    }
    if campo in campos_naturales and "#" in texto:
        texto = texto.replace("#", "Ñ")
        reparado = True
    danado = "�" in texto
    return texto, reparado, danado


def _digits(v: Any) -> str:
    s = _texto_limpio(v)
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return re.sub(r"\D", "", s)


def normalizar_telefono_argentina(valor: Any) -> tuple[str, str]:
    """Devuelve (telefono_10_digitos, motivo_error).

    Reglas EnvíosYA: sólo 10 dígitos, sin +54/54, sin 9 internacional,
    sin 0 interurbano y sin el 15 histórico cuando aparece como prefijo móvil.
    Nunca borra un '15' que forme parte de un número ya válido de 10 dígitos.
    """
    original = _digits(valor)
    if not original:
        return "", "Sin teléfono"

    d = original
    if d.startswith("0054"):
        d = d[4:]
    elif d.startswith("54") and len(d) >= 12:
        d = d[2:]

    # Formato internacional móvil: +54 9 AA ...
    if len(d) == 11 and d.startswith("9"):
        d = d[1:]

    # Prefijo interurbano 0. Ej.: 011 4149 2756.
    if len(d) == 11 and d.startswith("0"):
        d = d[1:]

    # Formato histórico: 0 + área + 15 + abonado, o área + 15 + abonado.
    # Sólo se evalúa si todavía NO son 10 dígitos.
    if len(d) != 10:
        sin_cero = d[1:] if d.startswith("0") else d
        candidatos = []
        for largo_area in (2, 3, 4):
            if len(sin_cero) == 12 and sin_cero[largo_area:largo_area + 2] == "15":
                candidatos.append(sin_cero[:largo_area] + sin_cero[largo_area + 2:])
        candidatos = [x for x in candidatos if len(x) == 10 and not x.startswith("0")]
        if len(candidatos) == 1:
            d = candidatos[0]

    if len(d) != 10:
        return "", f"No se pudo normalizar a 10 dígitos ({len(d)} dígitos)"
    if d.startswith("0"):
        return "", "El número normalizado no puede comenzar con 0"
    if len(set(d)) <= 2:
        return "", "Número sospechoso"
    return d, ""


def _es_email(v: Any) -> bool:
    s = _texto_limpio(v)
    return bool(re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", s))


def _es_patente(v: Any) -> bool:
    s = re.sub(r"[^A-Z0-9]", "", _norm_texto(v))
    return bool(re.fullmatch(r"(?:[A-Z]{3}\d{3}|[A-Z]{2}\d{3}[A-Z]{2}|\d{3}[A-Z]{3})", s))


def _es_anio(v: Any) -> bool:
    s = _digits(v)
    if not s:
        return False
    try:
        n = int(s)
    except Exception:
        return False
    return 1900 <= n <= datetime.now().year + 2


def _excel_serial_fecha(v: Any) -> date | None:
    try:
        n = float(v)
    except Exception:
        return None
    if 20_000 <= n <= 70_000:
        return date(1899, 12, 30) + timedelta(days=int(n))
    return None


def _parse_fecha(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    serial = _excel_serial_fecha(v)
    if serial:
        return serial
    s = _texto_limpio(v)
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _parece_dni(v: Any) -> bool:
    d = _digits(v)
    return 7 <= len(d) <= 8


def _parece_cp(v: Any) -> bool:
    s = _norm_texto(v).replace(" ", "")
    return bool(re.fullmatch(r"\d{4}", s) or re.fullmatch(r"[A-Z]\d{4}[A-Z]{3}", s))


def _parece_direccion(v: Any) -> bool:
    s = _norm_texto(v)
    return bool(re.search(r"\d", s) and re.search(r"[A-Z]", s) and len(s) >= 6)


def _parece_tipo(v: Any) -> bool:
    return _norm_texto(v) in TIPOS_CONOCIDOS


def _nombre_score(v: Any) -> float:
    s = _norm_texto(v)
    if not s or re.search(r"\d|@", s):
        return 0.0
    tokens = [x for x in re.split(r"\s+", s) if x]
    if not (2 <= len(tokens) <= 7):
        return 0.0
    comunes = sum(1 for t in tokens if t in NOMBRES_COMUNES)
    score = 0.25 + min(0.6, comunes * 0.3)
    if len(tokens) >= 3:
        score += 0.1
    return min(score, 1.0)


def separar_nombre_completo(valor: Any) -> tuple[str, str]:
    original = re.sub(r"\s+", " ", _texto_limpio(valor)).strip(" ,;")
    if not original:
        return "", ""
    if "," in original:
        izq, der = original.split(",", 1)
        return izq.strip(), der.strip()
    tokens = original.split()
    if len(tokens) == 1:
        return tokens[0], ""

    norm_tokens = [_norm_texto(x) for x in tokens]
    limite = None
    # Buscamos el primer nombre reconocible; permite apellidos compuestos.
    for i in range(1, len(tokens)):
        if norm_tokens[i] in NOMBRES_COMUNES:
            limite = i
            break
    if limite is None:
        limite = 1
    return " ".join(tokens[:limite]), " ".join(tokens[limite:])


def _detectar_compania_filename(nombre: str) -> str:
    n = _norm_texto(Path(nombre).stem).replace("_", " ").replace("-", " ")
    for clave, salida in COMPANIAS_FILENAME.items():
        if _norm_texto(clave) in n:
            return salida
    return ""


def _canon_header(v: Any) -> str:
    s = _norm_texto(v).lower()
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    return s


def _campo_por_header(v: Any) -> str | None:
    h = _canon_header(v)
    if not h:
        return None
    for campo, aliases in ALIASES_CAMPOS.items():
        for alias in aliases:
            if h == _canon_header(alias):
                return campo
    return None


def _fila_es_header(row: list[Any]) -> bool:
    detectados = [c for c in (_campo_por_header(v) for v in row) if c]
    fuertes = {"celular", "dni", "patente", "vencimiento", "email", "localidad", "nombre", "apellido", "nombre_completo"}
    return len(detectados) >= 2 or any(c in fuertes for c in detectados)


def _inferir_columnas(rows: list[list[Any]]) -> dict[int, str]:
    if not rows:
        return {}
    ancho = max(len(r) for r in rows)
    muestras = rows[:250]
    scores: dict[int, dict[str, float]] = {}
    for idx in range(ancho):
        vals = [r[idx] for r in muestras if idx < len(r) and _texto_limpio(r[idx])]
        if not vals:
            continue
        n = len(vals)
        metricas = {
            "email": sum(_es_email(v) for v in vals) / n,
            "celular": sum(bool(normalizar_telefono_argentina(v)[0]) for v in vals) / n,
            "patente": sum(_es_patente(v) for v in vals) / n,
            "anio": sum(_es_anio(v) for v in vals) / n,
            "fecha_origen": sum(bool(_parse_fecha(v)) for v in vals) / n,
            "dni": sum(_parece_dni(v) for v in vals) / n,
            "cp": sum(_parece_cp(v) for v in vals) / n,
            "direccion": sum(_parece_direccion(v) for v in vals) / n,
            "tipo": sum(_parece_tipo(v) for v in vals) / n,
            "nombre_completo": sum(_nombre_score(v) for v in vals) / n,
        }
        # Una columna de teléfono válida no debe terminar etiquetada DNI/CP.
        if metricas["celular"] >= 0.65:
            metricas["dni"] *= 0.1
            metricas["cp"] *= 0.1
        scores[idx] = metricas

    mapping: dict[int, str] = {}
    usados: set[str] = set()
    prioridad = ["email", "celular", "patente", "tipo", "fecha_origen", "anio", "cp", "dni", "direccion", "nombre_completo"]
    umbral = {"email": .55, "celular": .55, "patente": .45, "tipo": .45, "fecha_origen": .55, "anio": .55, "cp": .65, "dni": .65, "direccion": .55, "nombre_completo": .42}
    for campo in prioridad:
        candidatos = [(idx, sc.get(campo, 0.0)) for idx, sc in scores.items() if idx not in mapping]
        if candidatos:
            idx, puntaje = max(candidatos, key=lambda x: x[1])
            if puntaje >= umbral[campo] and campo not in usados:
                mapping[idx] = campo
                usados.add(campo)

    # Texto remanente: normalmente localidad en bases sin encabezado.
    for idx in range(ancho):
        if idx in mapping:
            continue
        vals = [_texto_limpio(r[idx]) for r in muestras if idx < len(r) and _texto_limpio(r[idx])]
        if not vals:
            continue
        alfab = [v for v in vals if not re.search(r"\d|@", v)]
        if len(alfab) / len(vals) >= .8:
            avg_tokens = sum(len(v.split()) for v in alfab) / len(alfab)
            if avg_tokens <= 3.0 and "localidad" not in usados:
                mapping[idx] = "localidad"
                usados.add("localidad")
                break
    return mapping


def _mapear_con_gemini(rows: list[list[Any]], mapping_actual: dict[int, str]) -> dict[int, str]:
    """Fallback muy acotado: sólo mapea columnas, jamás procesa 40k filas con IA."""
    if any(v == "celular" for v in mapping_actual.values()) and any(v in {"nombre", "nombre_completo"} for v in mapping_actual.values()):
        return mapping_actual
    try:
        from ai_gateway import begin_request, generate_with_fallback, obtener_cliente_gemini, DEFAULT_MODELS
        from google.genai import types
        cliente = obtener_cliente_gemini()
        if not cliente:
            return mapping_actual
        muestra = [[_texto_limpio(v)[:90] for v in r] for r in rows[:12]]
        prompt = (
            "Mapeá columnas de una base de contactos a estos campos canónicos: "
            "apellido,nombre,nombre_completo,dni,celular,localidad,cp,direccion,compania,patente,marca,modelo,anio,cliente,vencimiento,fecha_origen,email,tipo,poliza,provincia,pais. "
            "Respondé SOLO JSON con claves que sean índices de columna base 0 y valores campos canónicos. "
            "No inventes: si una columna es ambigua, omitila. Muestra:\n" + json.dumps(muestra, ensure_ascii=False)
        )
        begin_request()
        try:
            r, _modelo_usado = generate_with_fallback(
                client=cliente,
                models=DEFAULT_MODELS[:2],
                contents=prompt,
                config=types.GenerateContentConfig(response_mime_type="application/json", temperature=0),
                log_prefix="GEMINI /ENVIOS",
            )
            data = json.loads(r.text or "{}")
            out = dict(mapping_actual)
            validos = set(ALIASES_CAMPOS)
            for k, v in data.items():
                try:
                    i = int(k)
                except Exception:
                    continue
                if str(v) in validos and i not in out:
                    out[i] = str(v)
            return out
        except Exception:
            pass
    except Exception:
        pass
    return mapping_actual


def _leer_xlsx(datos: bytes) -> tuple[list[list[Any]], str]:
    wb = load_workbook(io.BytesIO(datos), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i > MAX_FILAS_POR_ARCHIVO:
                raise ValueError(f"La base supera el máximo de {MAX_FILAS_POR_ARCHIVO:,} filas por archivo.")
            vals = list(row)
            if any(_texto_limpio(v) for v in vals):
                rows.append(vals)
        return rows, ws.title
    finally:
        wb.close()


def _leer_csv(datos: bytes) -> tuple[list[list[Any]], str]:
    texto = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = datos.decode(enc)
            break
        except Exception:
            continue
    if texto is None:
        raise ValueError("No pude leer la codificación del CSV.")
    muestra = texto[:8192]
    try:
        dialect = csv.Sniffer().sniff(muestra, delimiters=",;\t|")
    except Exception:
        dialect = csv.excel
        dialect.delimiter = ";"
    rows = []
    for i, row in enumerate(csv.reader(io.StringIO(texto), dialect), start=1):
        if i > MAX_FILAS_POR_ARCHIVO:
            raise ValueError(f"La base supera el máximo de {MAX_FILAS_POR_ARCHIVO:,} filas por archivo.")
        if any(_texto_limpio(v) for v in row):
            rows.append(row)
    return rows, "CSV"


def _leer_archivo(nombre: str, datos: bytes) -> tuple[list[list[Any]], str]:
    ext = Path(nombre).suffix.lower()
    if ext in {".xlsx", ".xlsm"}:
        return _leer_xlsx(datos)
    if ext == ".csv":
        return _leer_csv(datos)
    if ext == ".xls":
        raise ValueError("El formato .xls antiguo no está soportado. Guardalo como .xlsx y volvé a subirlo.")
    raise ValueError("Formato no soportado. Usá Excel .xlsx/.xlsm o CSV.")


def _valor(row: list[Any], mapping: dict[int, str], campo: str) -> Any:
    for idx, c in mapping.items():
        if c == campo and idx < len(row):
            return row[idx]
    return ""


def _normalizar_registro(row: list[Any], mapping: dict[int, str], fuente: str, compania_fuente: str, fecha_modo: str = "conservar", usar_compania_fuente: bool = False) -> dict[str, Any]:
    nombre, rep_nombre, bad_nombre = _reparar_texto_fuente(_valor(row, mapping, "nombre"), "nombre")
    apellido, rep_apellido, bad_apellido = _reparar_texto_fuente(_valor(row, mapping, "apellido"), "apellido")
    nombre_completo, rep_completo, bad_completo = _reparar_texto_fuente(_valor(row, mapping, "nombre_completo"), "nombre_completo")
    if nombre_completo and not (nombre and apellido):
        ap2, no2 = separar_nombre_completo(nombre_completo)
        apellido = apellido or ap2
        nombre = nombre or no2

    tel_original = _texto_limpio(_valor(row, mapping, "celular"))
    tel, error_tel = normalizar_telefono_argentina(tel_original)

    fecha = _parse_fecha(_valor(row, mapping, "vencimiento"))
    vencimiento = fecha.strftime("%d/%m/%Y") if fecha else ""
    anio_raw = _valor(row, mapping, "anio")
    anio = ""
    if _es_anio(anio_raw):
        anio = str(int(float(_texto_limpio(anio_raw))))

    compania = _texto_limpio(_valor(row, mapping, "compania"))
    if not compania and usar_compania_fuente:
        compania = compania_fuente
    patente = re.sub(r"[^A-Za-z0-9]", "", _texto_limpio(_valor(row, mapping, "patente"))).upper()
    if patente and not _es_patente(patente):
        patente = _texto_limpio(_valor(row, mapping, "patente")).upper()

    fecha_origen_obj = _parse_fecha(_valor(row, mapping, "fecha_origen"))
    fecha_origen = fecha_origen_obj.strftime("%d/%m/%Y") if fecha_origen_obj else ""
    if not vencimiento and fecha_modo == "vencimiento" and fecha_origen:
        vencimiento = fecha_origen

    localidad, rep_localidad, bad_localidad = _reparar_texto_fuente(_valor(row, mapping, "localidad"), "localidad")
    direccion, rep_direccion, bad_direccion = _reparar_texto_fuente(_valor(row, mapping, "direccion"), "direccion")
    provincia, rep_provincia, bad_provincia = _reparar_texto_fuente(_valor(row, mapping, "provincia"), "provincia")
    marca, rep_marca, bad_marca = _reparar_texto_fuente(_valor(row, mapping, "marca"), "marca")
    modelo, rep_modelo, bad_modelo = _reparar_texto_fuente(_valor(row, mapping, "modelo"), "modelo")
    cliente, rep_cliente, bad_cliente = _reparar_texto_fuente(_valor(row, mapping, "cliente"), "cliente")
    tipo, rep_tipo, bad_tipo = _reparar_texto_fuente(_valor(row, mapping, "tipo"), "tipo")
    hubo_reparacion = any((rep_nombre, rep_apellido, rep_completo, rep_localidad, rep_direccion, rep_provincia, rep_marca, rep_modelo, rep_cliente, rep_tipo))
    texto_danado_critico = any((bad_nombre, bad_apellido, bad_completo, bad_localidad))
    texto_danado = any((bad_nombre, bad_apellido, bad_completo, bad_localidad, bad_direccion, bad_provincia, bad_marca, bad_modelo, bad_cliente, bad_tipo))

    return {
        "apellido": apellido,
        "nombre": nombre,
        "nombre_completo": nombre_completo or (f"{apellido} {nombre}".strip()),
        "dni": _digits(_valor(row, mapping, "dni")),
        "celular_original": tel_original,
        "celular": tel,
        "telefono_error": error_tel,
        "localidad": localidad,
        "cp": _texto_limpio(_valor(row, mapping, "cp")),
        "direccion": direccion,
        "compania": compania,
        "patente": patente,
        "marca": marca,
        "modelo": modelo,
        "anio": anio,
        "cliente": cliente,
        "vencimiento": vencimiento,
        "fecha_origen": fecha_origen,
        "email": _texto_limpio(_valor(row, mapping, "email")),
        "tipo": tipo,
        "poliza": _texto_limpio(_valor(row, mapping, "poliza")),
        "provincia": provincia,
        "pais": _texto_limpio(_valor(row, mapping, "pais")) or "Argentina",
        "fuente": fuente,
        "compania_fuente": compania_fuente,
        "caracteres_reparados": hubo_reparacion,
        "texto_danado": texto_danado,
        "texto_danado_critico": texto_danado_critico,
    }


def _score_riqueza(r: dict[str, Any]) -> int:
    campos = ("apellido", "nombre", "dni", "celular", "localidad", "cp", "direccion", "compania", "patente", "marca", "modelo", "anio", "vencimiento", "email", "tipo", "poliza", "provincia")
    return sum(1 for c in campos if _texto_limpio(r.get(c)))


def _limpiar_temporales() -> None:
    ahora = time.time()
    try:
        for p in TMP_DIR.glob("envios_*.*"):
            if ahora - p.stat().st_mtime > TTL_TEMP_SEGUNDOS:
                p.unlink(missing_ok=True)
    except Exception:
        pass


def _generar_excel_enviosya(registros: list[dict[str, Any]], token: str) -> Path:
    """Genera el XLSX sobre la plantilla real de EnvíosYA con memoria acotada.

    Se conservan título/encabezados/estructura del archivo original y se
    reescribe sólo sheetData. Las filas se escriben en streaming dentro del ZIP
    para que una base grande no haga caer al worker por memoria.
    """
    if not PLANTILLA_ENVIOSYA.exists():
        raise RuntimeError("Falta la plantilla oficial de EnvíosYA.")

    def cell_text(col: str, row_n: int, value: Any) -> str:
        text = _texto_limpio(value)
        if not text:
            return ""
        return f'<c t="inlineStr" r="{col}{row_n}"><is><t>{_xml_escape(text)}</t></is></c>'

    def cell_num(col: str, row_n: int, value: Any) -> str:
        d = _digits(value)
        if not d:
            return ""
        return f'<c r="{col}{row_n}" s="65"><v>{int(d)}</v></c>'

    salida = TMP_DIR / f"envios_{token}.xlsx"
    with zipfile.ZipFile(PLANTILLA_ENVIOSYA, "r") as zin, zipfile.ZipFile(salida, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename != "xl/worksheets/sheet1.xml":
                zout.writestr(item, zin.read(item.filename))
                continue

            original = zin.read(item.filename).decode("utf-8")
            m = re.search(r'(<sheetData[^>]*>)(.*?)(</sheetData>)', original, re.S)
            if not m:
                raise RuntimeError("La plantilla de EnvíosYA no contiene sheetData.")
            filas_base = []
            for rm in re.finditer(r'<row\b[^>]*\br="(\d+)"[^>]*>.*?</row>', m.group(2), re.S):
                if int(rm.group(1)) < 3:
                    filas_base.append(rm.group(0))
            prefix = original[:m.start()] + m.group(1) + "".join(filas_base)
            suffix = m.group(3) + original[m.end():]

            with zout.open(item, "w") as out:
                out.write(prefix.encode("utf-8"))
                for row_n, r in enumerate(registros, start=3):
                    cells = [
                        cell_text("B", row_n, r.get("apellido")),
                        cell_text("C", row_n, r.get("nombre")),
                        cell_num("D", row_n, r.get("dni")),
                        cell_num("E", row_n, r.get("celular")),
                        cell_text("F", row_n, r.get("localidad")),
                        cell_text("G", row_n, r.get("compania")),
                        cell_text("H", row_n, r.get("patente")),
                        cell_text("I", row_n, r.get("marca")),
                        cell_text("J", row_n, r.get("modelo")),
                        cell_num("K", row_n, r.get("anio")),
                        cell_text("L", row_n, r.get("cliente")),
                        cell_text("M", row_n, r.get("vencimiento")),
                    ]
                    fila = f'<row r="{row_n}">{"".join(cells)}</row>'
                    out.write(fila.encode("utf-8"))
                out.write(suffix.encode("utf-8"))
    return salida


def _generar_csv_notificaciones(registros: list[dict[str, Any]], token: str) -> tuple[Path, int]:
    """CSV de importación de notificaciones según el orden exigido por EnvíosYA.

    apellido,nombre,celular,fecha,XXX,XXX,XXX,patente,XXX,poliza
    Los XXX son posiciones reservadas y se escriben literalmente.
    """
    salida = TMP_DIR / f"envios_notificaciones_{token}.csv"
    aptos = [r for r in registros if r.get("apellido") and r.get("nombre") and r.get("celular") and r.get("vencimiento")]
    with salida.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=",", lineterminator="\r\n")
        for r in aptos:
            w.writerow([
                r.get("apellido", ""), r.get("nombre", ""), r.get("celular", ""),
                r.get("vencimiento", ""), "XXX", "XXX", "XXX", r.get("patente", ""),
                "XXX", r.get("poliza", ""),
            ])
    return salida, len(aptos)


def _generar_excel_maestro(registros: list[dict[str, Any]], token: str) -> Path:
    salida = TMP_DIR / f"envios_maestro_{token}.xlsx"
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.title = "Base normalizada"
    headers = [
        "Apellido", "Nombre", "Nombre original", "DNI", "Celular original", "Celular EnvíosYA",
        "Email", "Dirección", "CP", "Localidad", "Provincia", "País", "Compañía", "Patente",
        "Marca", "Modelo", "Año", "Póliza", "Vencimiento", "Fecha origen", "Tipo", "Fuente",
        "Compañía sugerida por fuente", "Estado teléfono", "Caracteres reparados", "Observaciones",
    ]
    ws.append(headers)
    for r in registros:
        ws.append([
            r.get("apellido",""), r.get("nombre",""), r.get("nombre_completo",""), r.get("dni",""),
            r.get("celular_original",""), r.get("celular",""), r.get("email",""), r.get("direccion",""),
            r.get("cp",""), r.get("localidad",""), r.get("provincia",""), r.get("pais",""), r.get("compania",""),
            r.get("patente",""), r.get("marca",""), r.get("modelo",""), r.get("anio",""), r.get("poliza",""),
            r.get("vencimiento",""), r.get("fecha_origen",""), r.get("tipo",""), r.get("fuente",""),
            r.get("compania_fuente",""), "VALIDO" if r.get("celular") else "REVISAR",
            "SI" if r.get("caracteres_reparados") else "",
            (r.get("telefono_error", "") + (" | Texto de origen con carácter ilegible (�)" if r.get("texto_danado") else "")).strip(" |"),
        ])
    wb.save(salida)
    wb.close()
    return salida


def procesar_bases(archivos: list[tuple[str, bytes]], fecha_modo: str = "conservar", usar_compania_fuente: bool = False) -> dict[str, Any]:
    _limpiar_temporales()
    if not archivos:
        raise ValueError("Seleccioná al menos una base.")
    if len(archivos) > MAX_ARCHIVOS:
        raise ValueError(f"Podés procesar hasta {MAX_ARCHIVOS} archivos por lote.")

    todos: list[dict[str, Any]] = []
    info_archivos = []
    total_entrada = 0
    for nombre, datos in archivos:
        rows, hoja = _leer_archivo(nombre, datos)
        if not rows:
            info_archivos.append({"nombre": nombre, "filas": 0, "estado": "vacío", "columnas": {}})
            continue
        tiene_header = _fila_es_header(rows[0])
        if tiene_header:
            headers = rows[0]
            mapping = {i: c for i, v in enumerate(headers) if (c := _campo_por_header(v))}
            data_rows = rows[1:]
        else:
            headers = []
            data_rows = rows
            mapping = _inferir_columnas(data_rows)
        mapping = _mapear_con_gemini(data_rows, mapping)
        compania_fuente = _detectar_compania_filename(nombre)
        total_entrada += len(data_rows)
        for row in data_rows:
            reg = _normalizar_registro(row, mapping, nombre, compania_fuente, fecha_modo=fecha_modo, usar_compania_fuente=usar_compania_fuente)
            # Ignoramos filas sin ningún dato de persona/contacto.
            if not any(reg.get(k) for k in ("nombre_completo", "celular_original", "dni", "email", "localidad")):
                continue
            todos.append(reg)
        columnas = {str(i + 1): campo for i, campo in sorted(mapping.items())}
        info_archivos.append({
            "nombre": nombre,
            "hoja": hoja,
            "filas": len(data_rows),
            "encabezados": bool(tiene_header),
            "columnas": columnas,
            "compania_detectada": compania_fuente,
        })

    validos_por_tel: dict[str, dict[str, Any]] = {}
    invalidos = []
    duplicados = 0
    for r in todos:
        tel = r.get("celular") or ""
        if not tel or r.get("texto_danado_critico"):
            if r.get("texto_danado_critico") and not r.get("telefono_error"):
                r["telefono_error"] = "Texto principal del origen contiene un carácter ilegible (�)"
            invalidos.append(r)
            continue
        if tel in validos_por_tel:
            duplicados += 1
            actual = validos_por_tel[tel]
            if _score_riqueza(r) > _score_riqueza(actual):
                validos_por_tel[tel] = r
        else:
            validos_por_tel[tel] = r

    validos = list(validos_por_tel.values())
    validos.sort(key=lambda r: (_norm_texto(r.get("localidad")), _norm_texto(r.get("apellido")), _norm_texto(r.get("nombre"))))
    token = uuid.uuid4().hex
    salida = _generar_excel_enviosya(validos, token)
    salida_notif, notificaciones_aptas = _generar_csv_notificaciones(validos, token)
    salida_maestro = _generar_excel_maestro(todos, token)

    preview = []
    for r in validos[:120]:
        preview.append({**{k: r.get(k, "") for k in ("apellido", "nombre", "celular", "email", "localidad", "cp", "compania", "patente", "tipo", "fuente")}, "estado": "VALIDO", "motivo": ""})
    for r in invalidos[:80]:
        preview.append({**{k: r.get(k, "") for k in ("apellido", "nombre", "celular_original", "email", "localidad", "cp", "compania", "patente", "tipo", "fuente")}, "celular": r.get("celular_original", ""), "estado": "REVISAR", "motivo": r.get("telefono_error") or "Teléfono inválido"})

    return {
        "token": token,
        "archivo": salida.name,
        "resumen": {
            "filas_entrada": total_entrada,
            "contactos_detectados": len(todos),
            "exportables": len(validos),
            "revisar": len(invalidos),
            "duplicados": duplicados,
            "notificaciones_aptas": notificaciones_aptas,
            "con_email": sum(1 for r in validos if _es_email(r.get("email"))),
            "con_vencimiento": sum(1 for r in validos if r.get("vencimiento")),
            "caracteres_reparados": sum(1 for r in todos if r.get("caracteres_reparados")),
            "texto_danado": sum(1 for r in todos if r.get("texto_danado")),
        },
        "archivos": info_archivos,
        "preview": preview,
        "fecha_modo": fecha_modo,
        "notificaciones_disponibles": notificaciones_aptas > 0,
    }


def obtener_exportacion(token: str, tipo: str = "contactos") -> Path | None:
    if not re.fullmatch(r"[a-f0-9]{32}", str(token or "")):
        return None
    nombres = {
        "contactos": f"envios_{token}.xlsx",
        "notificaciones": f"envios_notificaciones_{token}.csv",
        "maestro": f"envios_maestro_{token}.xlsx",
    }
    nombre = nombres.get(tipo)
    if not nombre:
        return None
    p = TMP_DIR / nombre
    if not p.exists():
        return None
    if time.time() - p.stat().st_mtime > TTL_TEMP_SEGUNDOS:
        p.unlink(missing_ok=True)
        return None
    return p


def obtener_excel(token: str) -> Path | None:
    return obtener_exportacion(token, "contactos")
