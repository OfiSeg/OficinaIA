import re
import os
import time
import json
from datetime import datetime, date
from pathlib import Path
from google.genai import types
from ai_gateway import obtener_cliente_gemini, generate_with_fallback, DEFAULT_MODELS
from openpyxl import load_workbook
from companias import normalizar_compania, aliases_companias
from context_router import construir_plan_base, es_consulta_comparativa
from sofia_prompt import build_sofia_prompt
from sofia_tools import TOOL_DEFINITIONS
from document_search import buscar_en_documentos
from metadata_store import cargar_metadatos
from excel_analytics import analizar as analizar_dataset_excel

try:
    from storage_r2 import descargar_excel_interno, EXCEL_INTERNO_R2_KEY
except Exception:
    descargar_excel_interno = None
    EXCEL_INTERNO_R2_KEY = "excel_interno.xlsx"


# ==========================================================
# CONFIGURACIÓN
# ==========================================================




# Cache por proceso. OficinaIA mantiene 1 worker mientras el Excel/R2 sea la
# fuente de verdad; con múltiples workers este cache NO sería compartido.
_CACHE_EXCEL = {"datos": None, "cargado_en": 0.0}
TTL_CACHE_EXCEL_SEGUNDOS = 10

BASE_DIR = Path(__file__).resolve().parent
EXCEL_INTERNO = BASE_DIR / "excel_interno.xlsx"


# ==========================================================
# GEMINI
# ==========================================================


def _texto_fila(fila):
    return " ".join(str(v or "") for v in fila.values()).strip()

def _coincidencia_identificador(pregunta, fila):
    """Devuelve True si la consulta contiene un identificador fuerte de esa fila."""
    q = _normalizar_texto(pregunta)
    for campo in ("PATENTE", "NUMERO", "NRO", "POLIZA", "PÓLIZA"):
        valor = str(fila.get(campo, "")).strip()
        if valor and _normalizar_texto(valor) in q:
            return True
    return False

def _asegurar_excel_local_para_ia():
    """Asegura el Excel interno para la IA usando la misma fuente que la UI (P0.6).

    Si R2 está configurado, SIEMPRE descarga de R2 (aunque exista copia local
    del repo). Así Gemini no trabaja con el xlsx versionado en git mientras
    la grilla de /notas ya tiene la versión de nube.
    """
    if descargar_excel_interno is None:
        return EXCEL_INTERNO.exists()
    try:
        ok = bool(descargar_excel_interno(EXCEL_INTERNO, EXCEL_INTERNO_R2_KEY))
        if ok:
            return True
        return EXCEL_INTERNO.exists()
    except Exception as error:
        print("ERROR RECUPERANDO EXCEL INTERNO PARA IA:", error)
        return EXCEL_INTERNO.exists()


def invalidar_cache_excel_interno():
    """Invalida la copia en memoria después de una escritura confirmada."""
    _CACHE_EXCEL["datos"] = None
    _CACHE_EXCEL["cargado_en"] = 0.0


def _cargar_excel_interno():
    ahora = time.monotonic()
    cache = _CACHE_EXCEL.get("datos")
    if cache is not None and (ahora - _CACHE_EXCEL["cargado_en"]) < TTL_CACHE_EXCEL_SEGUNDOS:
        return cache

    if not _asegurar_excel_local_para_ia():
        return []
    try:
        wb = load_workbook(EXCEL_INTERNO, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            _CACHE_EXCEL["datos"] = []
            _CACHE_EXCEL["cargado_en"] = ahora
            return []
        headers = [str(x or "").strip().upper() for x in rows[0]]
        datos = []
        for row in rows[1:]:
            fila = {
                headers[i] if i < len(headers) and headers[i] else f"COLUMNA_{i+1}":
                str(row[i] or "").strip()
                for i in range(min(len(headers), len(row)))
            }
            if any(fila.values()):
                datos.append(fila)
        wb.close()
        _CACHE_EXCEL["datos"] = datos
        _CACHE_EXCEL["cargado_en"] = time.monotonic()
        print("EXCEL INTERNO IA:", len(datos), "registros cargados.")
        return datos
    except Exception as error:
        print("ERROR EXCEL INTERNO:", error)
        return []


def _buscar_en_registros(pregunta, datos, etiqueta):
    if not datos:
        return []

    q = _normalizar_texto(pregunta)
    palabras = [p for p in re.findall(r"[a-z0-9]+", q) if len(p) >= 3]
    palabras = list(_expandir_sinonimos_tokens(palabras))
    cliente_exacto = buscar_cliente_exactamente(pregunta, datos) if any("CLIENTE" in f for f in datos) else None

    # Identificador fuerte: restringimos el contexto únicamente a las filas
    # que contienen ese identificador. Esto evita mezclar pólizas/patentes.
    identificadas = [f for f in datos if _coincidencia_identificador(pregunta, f)]
    if identificadas:
        return [(1000, f) for f in identificadas]

    if cliente_exacto:
        cliente_norm = _normalizar_texto(cliente_exacto)
        filas_cliente = [
            f for f in datos
            if _normalizar_texto(f.get("CLIENTE", "")) == cliente_norm
        ]
        if filas_cliente:
            return [(900, f) for f in filas_cliente]

    resultados = []
    for fila in datos:
        texto = _normalizar_texto(_texto_fila(fila))
        coincidencias = sum(1 for palabra in palabras if palabra in texto)
        if coincidencias:
            resultados.append((coincidencias, fila))
    # Si la consulta nombra una compañía, devolvemos TODAS sus filas.
    # Esto es imprescindible para conteos, listados y agrupaciones: un top-N
    # semántico no puede representar el dataset completo.
    companias = _companias_mencionadas(pregunta, datos)
    if companias:
        campos_cia = _campos_compania(datos)
        filas_compania = [
            f for f in datos
            if any(
                _normalizar_texto(f.get(campo, "")) in companias
                for campo in campos_cia
            )
        ]
        if filas_compania:
            return [(50, f) for f in filas_compania]

    resultados.sort(key=lambda x: x[0], reverse=True)
    return resultados[:50]


# ==========================================================
# CONSULTAS ESTRUCTURADAS SOBRE EL DATASET COMPLETO
# ==========================================================

_ALIAS_CIAS = {
    alias: codigo_display[0]
    for alias, codigo_display in aliases_companias().items()
}



def _identidad_unica(fila):
    # Conserva la lógica original: primero identificadores fuertes y luego
    # cliente/asegurado/nombre y, como último recurso, póliza.
    for aliases in (
        ("DNI", "DOCUMENTO", "CUIT", "CUIL"),
        ("CLIENTE", "ASEGURADO", "NOMBRE"),
        ("POLIZA", "PÓLIZA"),
    ):
        clave = _campo_por_alias(fila, aliases)
        if clave and str(fila.get(clave, "")).strip():
            return _normalizar_texto(fila.get(clave, ""))
    return None


def _deduplicar_personas(filas):
    """Deduplica sólo cuando hay un identificador razonablemente confiable."""
    if not filas:
        return []
    ids = []
    sin_id = []
    vistos = set()
    for fila in filas:
        ident = _identidad_unica(fila)
        if ident:
            if ident in vistos:
                continue
            vistos.add(ident)
            ids.append(fila)
        else:
            sin_id.append(fila)
    return ids + sin_id


def _campo_identidad_principal(datos):
    if not datos:
        return None
    aliases = (
        "DNI", "DOCUMENTO", "CUIT", "CUIL",
        "CLIENTE", "ASEGURADO", "NOMBRE",
    )
    for fila in datos[: min(len(datos), 20)]:
        campo = _campo_por_alias(fila, aliases)
        if campo and any(str(f.get(campo, "")).strip() for f in datos):
            return campo
    return None


def _campos_compania(datos):
    if not datos:
        return []
    candidatos = ("CIA", "COMPAÑIA", "COMPANIA", "COMPAÑÍA", "ASEGURADORA", "COMPANIA DE SEGUROS")
    presentes = set()
    for fila in datos[: min(len(datos), 3)]:
        presentes.update(str(k).strip().upper() for k in fila.keys())
    return [c for c in candidatos if c in presentes]


def _companias_mencionadas(pregunta, datos):
    q = _normalizar_texto(pregunta)
    mencionadas = set()
    for alias, canon in _ALIAS_CIAS.items():
        if re.search(rf"\b{re.escape(_normalizar_texto(alias))}\b", q):
            mencionadas.add(_normalizar_texto(normalizar_compania(canon)))
    return mencionadas


def _campo_por_alias(fila, aliases):
    for alias in aliases:
        for clave in fila.keys():
            if _normalizar_texto(clave) == _normalizar_texto(alias):
                return clave
    return None


def _filas_por_companias(pregunta, datos):
    mencionadas = _companias_mencionadas(pregunta, datos)
    campos = _campos_compania(datos)
    if not mencionadas or not campos:
        return datos
    salida = []
    for fila in datos:
        valor = next((fila.get(c, "") for c in campos), "")
        if _normalizar_texto(valor) in mencionadas:
            salida.append(fila)
    return salida



def buscar_en_excel_interno(pregunta):
    datos = _cargar_excel_interno()
    resultados = _buscar_en_registros(pregunta, datos, "Excel interno")
    if not resultados:
        return ""

    contexto = "FUENTE: Excel interno de OficinaIA\n"
    for _, fila in resultados:
        contexto += "REGISTRO INTERNO:\n"
        for campo, valor in fila.items():
            contexto += f"{campo}: {valor}\n"
        contexto += "\n--------------------\n"
    return contexto


def _normalizar_texto(texto):
    import unicodedata
    texto = str(texto or '').lower().strip()
    texto = unicodedata.normalize('NFKD', texto)
    return ''.join(c for c in texto if not unicodedata.combining(c))


# ==========================================================
# SINÓNIMOS DE DOMINIO (seguros)
# ==========================================================
# El dataset no siempre usa la misma palabra que el usuario ("grúa" en la
# pregunta, "remolque" en la planilla). Sin esto, la primera búsqueda daba 0
# resultados y el prompt forzaba un segundo (y tercer...) intento en el mismo
# request, lo que terminaba en timeout (502) o en una cadena de errores (500).
# Expandiendo la CONSULTA con sinónimos, la primera búsqueda ya encuentra la
# fila correcta la gran mayoría de las veces.
_SINONIMOS_DOMINIO = [
    {"grua", "gruas", "remolque", "remolques", "auxilio", "traslado", "arrastre", "acarreo"},
    {"asistencia", "asistencias", "sat", "auxilio", "socorro"},
    {"choque", "choques", "colision", "colisiones", "siniestro", "siniestros", "accidente", "accidentes"},
    {"robo", "robos", "hurto", "hurtos"},
    {"cristales", "cristal", "vidrios", "vidrio", "parabrisas", "lunas"},
    {"franquicia", "franquicias", "deducible", "deducibles"},
    {"poliza", "polizas"},
    {"vencimiento", "vencimientos", "renovacion", "renovaciones"},
    {"vehiculo", "vehiculos", "auto", "autos", "unidad", "unidades", "rodado", "rodados"},
]


def _expandir_sinonimos_tokens(tokens):
    """Devuelve el conjunto de tokens original + sinónimos conocidos."""
    base = {t for t in tokens if t}
    expandido = set(base)
    for grupo in _SINONIMOS_DOMINIO:
        if base & grupo:
            expandido |= grupo
    return expandido



# ==========================================================
# CONSULTAR GEMINI
# ==========================================================

def _tokens_relevancia(texto):
    return [p for p in re.findall(r"[a-z0-9]+", _normalizar_texto(texto)) if len(p) >= 3]


def _puntuar_fuente(pregunta, contenido, prioridad=0):
    """Puntúa una fuente por relevancia, sin confundir cantidad con calidad."""
    q = _normalizar_texto(pregunta)
    tokens = set(_tokens_relevancia(q))
    texto = _normalizar_texto(contenido)
    if not tokens or not texto:
        return prioridad

    score = prioridad
    score += sum(1 for token in tokens if token in texto)

    # Identificadores concretos pesan mucho más que coincidencias generales.
    for patron in (
        r"\b[a-z]{2,4}\d{2,6}\b",          # patente/código
        r"\b\d{5,12}\b",                   # número/póliza
    ):
        for identificador in re.findall(patron, q):
            if identificador in texto:
                score += 30

    if len(q) >= 12 and q in texto:
        score += 25
    return score


def _documentos_desde_contexto(contexto):
    """
    Convierte los fragmentos recuperados de PDFs en fuentes individuales.
    Así Gemini no recibe 7 documentos sólo porque todos tuvieron alguna
    coincidencia: cada archivo compite como una fuente propia.
    """
    fuentes = {}

    # PDF adjunto: es la fuente explícita que el usuario puso en el chat.
    adjunto = re.search(
        r"===== PDF ADJUNTADO EN EL CHAT =====(.*?)===== FIN PDF ADJUNTADO =====",
        contexto or "",
        flags=re.S,
    )
    if adjunto:
        bloque = adjunto.group(1).strip()
        m = re.search(r"ARCHIVO:\s*([^\n]+)", bloque)
        nombre = m.group(1).strip() if m else "PDF adjunto"
        fuentes[f"PDF adjunto \"{nombre}\""] = bloque

    # Fragmentos de documentación normal.
    bloques = re.findall(
        r"===== FRAGMENTO DE DOCUMENTO =====(.*?)(?=^===== FRAGMENTO DE DOCUMENTO =====|\Z)",
        contexto or "",
        flags=re.S | re.M,
    )
    for bloque in bloques:
        m = re.search(r"ARCHIVO:\s*([^\n]+)", bloque)
        if not m:
            continue
        nombre = m.group(1).strip()
        clave = f'Manual/documento "{nombre}"'
        fuentes[clave] = fuentes.get(clave, "") + "\n" + bloque.strip()

    return fuentes


def _agregar_fuentes(respuesta, fuentes, uso_web=False):
    texto = str(respuesta or "").strip()
    if not texto:
        return texto

    # El modelo ya recibe instrucciones de citar las fuentes seleccionadas.
    # Este respaldo sólo agrega las que realmente se seleccionaron, nunca
    # todas las fuentes consultadas durante el ranking.
    existentes = _normalizar_texto(texto)
    faltantes = [
        fuente for fuente in fuentes
        if _normalizar_texto(fuente) not in existentes
    ]

    if uso_web and not any(
        palabra in existentes
        for palabra in ("internet", "informacion publica", "sitio oficial")
    ):
        if len(fuentes) < 2:
            faltantes.append("Información pública de Internet")

    if not faltantes:
        return texto
    return texto + "\n\n**Fuentes:** " + " · ".join(faltantes)


# ==========================================================
# FUNCTION CALLING DE GEMINI
# ==========================================================
# V20 Etapa 3: los contratos de herramientas viven en sofia_tools.py.
# Este módulo conserva únicamente su ejecución y la lógica de dominio.


def _dataset_estructurado():
    return _cargar_excel_interno(), "Excel interno"


def _valor_campo(fila, campo):
    if not campo:
        return None
    clave = _campo_por_alias(fila, (campo,))
    return fila.get(clave, "") if clave else None


def _parsear_fecha_excel(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor or '').strip()
    if not texto:
        return None
    # Cubre fecha real serializada por openpyxl y los formatos históricos.
    for fmt in (
        '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d',
        '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S',
    ):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            pass
    return None


def _filtrar_filas(
    filas, compania=None, campo=None, valor=None,
    tipo_vehiculo=None, desde=None, hasta=None, campo_fecha=None,
):
    salida = list(filas)
    filtros = {}

    if compania:
        objetivo = _normalizar_texto(normalizar_compania(compania))
        campos = []
        for f in salida[:20]:
            for alias in ('CIA', 'COMPAÑIA', 'COMPANIA', 'COMPAÑÍA', 'ASEGURADORA', 'COMPANIA DE SEGUROS'):
                c = _campo_por_alias(f, (alias,))
                if c and c not in campos:
                    campos.append(c)
        salida = [
            f for f in salida
            if any(_normalizar_texto(normalizar_compania(f.get(c, ''))) == objetivo for c in campos)
        ]
        filtros['compania'] = normalizar_compania(compania)

    if tipo_vehiculo:
        objetivo = _normalizar_texto(tipo_vehiculo)
        equivalencias_tipo = {
            'remolque': {'remolque', 'trailer'},
            'remolques': {'remolque', 'trailer'},
            'trailer': {'remolque', 'trailer'},
            'trailers': {'remolque', 'trailer'},
        }
        objetivos = equivalencias_tipo.get(objetivo, {objetivo})
        aliases_vehiculo = ('VEHICULO', 'VEHÍCULO', 'TIPO VEHICULO', 'TIPO DE VEHICULO', 'MARCA MODELO')
        def coincide_tipo(fila):
            clave = _campo_por_alias(fila, aliases_vehiculo)
            texto_vehiculo = _normalizar_texto(fila.get(clave, '') if clave else '')
            return any(token in texto_vehiculo for token in objetivos)
        salida = [f for f in salida if coincide_tipo(f)]
        filtros['tipo_vehiculo'] = str(tipo_vehiculo).strip()
        if len(objetivos) > 1:
            filtros['sinonimos_tipo_vehiculo'] = sorted(objetivos)

    if campo and valor is not None:
        objetivo = _normalizar_texto(valor)
        salida = [f for f in salida if objetivo in _normalizar_texto(_valor_campo(f, campo) or '')]
        filtros['campo'] = str(campo).strip()
        filtros['valor'] = str(valor).strip()

    if desde or hasta:
        desde_fecha = _parsear_fecha_excel(desde) if desde else None
        hasta_fecha = _parsear_fecha_excel(hasta) if hasta else None
        if desde and desde_fecha is None:
            raise ValueError('Fecha desde inválida. Usá DD/MM/AAAA.')
        if hasta and hasta_fecha is None:
            raise ValueError('Fecha hasta inválida. Usá DD/MM/AAAA.')
        aliases_fecha = (campo_fecha,) if campo_fecha else ('EMITIDO DÍA:', 'EMITIDO DIA', 'EMITIDO', 'FECHA EMISION', 'FECHA DE EMISION')
        filtradas = []
        for f in salida:
            clave_fecha = _campo_por_alias(f, aliases_fecha)
            fecha = _parsear_fecha_excel(f.get(clave_fecha, '') if clave_fecha else '')
            if fecha is None:
                continue
            if desde_fecha and fecha < desde_fecha:
                continue
            if hasta_fecha and fecha > hasta_fecha:
                continue
            filtradas.append(f)
        salida = filtradas
        if desde_fecha:
            filtros['desde'] = desde_fecha.strftime('%d/%m/%Y')
        if hasta_fecha:
            filtros['hasta'] = hasta_fecha.strftime('%d/%m/%Y')
        filtros['campo_fecha'] = campo_fecha or 'EMITIDO DÍA:'

    return salida, filtros


def consultar_excel(pregunta_o_filtro):
    """Herramienta de búsqueda estructurada; no decide intención."""
    interno = _cargar_excel_interno()
    datos, fuente = _dataset_estructurado()
    if not datos:
        return {"fuente": fuente, "registros": [], "cantidad": 0}

    q = _normalizar_texto(pregunta_o_filtro)
    identificadas = [f for f in datos if _coincidencia_identificador(pregunta_o_filtro, f)]
    if identificadas:
        filas = identificadas
    else:
        palabras = [p for p in re.findall(r"[a-z0-9]+", q) if len(p) >= 3]
        puntuadas = []
        for fila in datos:
            texto = _normalizar_texto(_texto_fila(fila))
            score = sum(1 for palabra in palabras if palabra in texto)
            if score:
                puntuadas.append((score, fila))
        puntuadas.sort(key=lambda x: x[0], reverse=True)
        filas = [fila for _, fila in puntuadas[:100]]

    return {
        "fuente": fuente,
        "cantidad": len(filas),
        "registros": filas,
    }


def contar_registros(
    compania=None, campo=None, valor=None, tipo_vehiculo=None,
    desde=None, hasta=None, campo_fecha=None, tipo_conteo='filas',
):
    """Conteo determinístico sobre el dataset COMPLETO, nunca sobre previews."""
    datos, fuente = _dataset_estructurado()
    filas, filtros = _filtrar_filas(
        datos, compania=compania, campo=campo, valor=valor,
        tipo_vehiculo=tipo_vehiculo, desde=desde, hasta=hasta, campo_fecha=campo_fecha,
    )
    campo_identidad = _campo_identidad_principal(filas)
    filas_unicas = _deduplicar_personas(filas) if campo_identidad else filas
    tipo = _normalizar_texto(tipo_conteo or 'filas')
    cantidad = len(filas_unicas) if tipo in {'unicos', 'unico', 'personas', 'asegurados'} else len(filas)
    return {
        'fuente': fuente,
        'cantidad': cantidad,
        'tipo_conteo': 'unicos' if tipo in {'unicos', 'unico', 'personas', 'asegurados'} else 'filas',
        'total_filas': len(filas),
        'total_unicos': len(filas_unicas),
        'campo_identidad': campo_identidad,
        'filtros_aplicados': filtros,
        'dataset_total_filas': len(datos),
    }


def analizar_excel(consulta=None, operacion=None, campo=None, agrupar_por=None, compania=None, valor=None, excluir_valor=None, limite=None):
    """Analítica determinística sobre TODO el Excel interno.

    Resuelve agrupaciones, rankings, porcentajes, duplicados, vacíos y
    clasificación auto/moto sin contar previews ni pedirle a Gemini que haga
    cálculos sobre muestras.
    """
    datos, fuente = _dataset_estructurado()
    resultado = analizar_dataset_excel(
        datos, consulta=consulta, operacion=operacion, campo=campo,
        agrupar_por=agrupar_por, compania=compania, valor=valor,
        excluir_valor=excluir_valor, limite=limite,
    )
    if isinstance(resultado, dict):
        resultado = dict(resultado)
        resultado.setdefault("fuente", fuente)
        resultado.setdefault("dataset_total_filas", len(datos))
    return resultado


def proponer_registro_excel(campos):
    """
    Prepara una propuesta de alta; nunca escribe directamente en el Excel.

    Las claves se normalizan contra el esquema real de la planilla. Los campos
    faltantes se conservan vacíos para que el usuario pueda completarlos en la
    confirmación, en lugar de desaparecer silenciosamente.
    """
    campos_validos = (
        "ASEGURADO",
        "NUMERO",
        "VEHICULO",
        "PATENTE",
        "ENVIOS YA",
        "CIA",
        "MEDIO DE PAGO",
        "CP",
        "MAIL",
        "TELEFONO",
    )
    if not isinstance(campos, dict):
        return {"propuesta": {}, "valida": False}

    propuesta = {
        campo: str(campos.get(campo, "") or "").strip()
        for campo in campos_validos
    }

    # Nunca aceptar claves inventadas por el modelo.
    for clave, valor in campos.items():
        clave_norm = _normalizar_texto(clave)
        for campo in campos_validos:
            if _normalizar_texto(campo) == clave_norm:
                propuesta[campo] = str(valor or "").strip()
                break

    propuesta["CIA"] = normalizar_compania(propuesta.get("CIA", ""))

    tiene_asegurado = bool(propuesta["ASEGURADO"])
    tiene_identificador = bool(propuesta["NUMERO"] or propuesta["PATENTE"])

    return {
        "propuesta": propuesta,
        "valida": bool(tiene_asegurado and tiene_identificador),
        "faltantes_minimos": [
            campo for campo, ok in (
                ("ASEGURADO", tiene_asegurado),
                ("NUMERO o PATENTE", tiene_identificador),
            ) if not ok
        ],
    }


def guardar_metadato_relevante(titulo, contenido):
    """
    Prepara una propuesta de metadato reutilizable; nunca escribe directamente.

    Sólo se aceptan datos objetivos, relativamente estables y reutilizables
    en consultas futuras: definiciones operativas, procedimientos, cifras,
    cantidades, límites o condiciones de seguros. No hace falta que el dato
    pertenezca a una compañía concreta. Se reutiliza la misma recuperación
    existente para evitar proponer duplicados obvios.
    """
    titulo = str(titulo or "").strip()
    contenido = str(contenido or "").strip()

    if not titulo or not contenido:
        return {"propuesta": None, "valida": False}

    if len(titulo) > 200:
        titulo = titulo[:200]

    # No guardar conversaciones, opiniones, instrucciones temporales ni
    # texto demasiado largo. El metadato debe ser una ficha puntual y
    # reutilizable, aunque sea una definición general del trabajo de seguros.
    if len(contenido) > 1200:
        contenido = contenido[:1200].rsplit(" ", 1)[0].strip()

    texto = _normalizar_texto(f"{titulo} {contenido}")
    patrones_descartables = (
        "creo que", "me parece", "quizas", "quiza", "podrias",
        "te recomiendo", "como puedo", "que opinas",
    )
    if any(patron in texto for patron in patrones_descartables):
        return {"propuesta": None, "valida": False, "motivo": "dato_no_objetivo"}

    try:
        existentes = _cargar_metadatos()
        for ficha in existentes:
            if _normalizar_texto(contenido) == _normalizar_texto(ficha.get("contenido", "")):
                return {
                    "propuesta": None,
                    "valida": False,
                    "duplicado": True,
                    "metadato_existente": {
                        "id": ficha.get("id"),
                        "titulo": ficha.get("titulo"),
                    },
                }
    except Exception as error:
        print("ERROR VALIDANDO METADATO PROPUESTO:", error)

    return {
        "propuesta": {
            "titulo": titulo,
            "contenido": contenido,
        },
        "valida": True,
    }



def buscar_en_manuales(consulta):
    """Consulta Manuales/Pólizas sin depender de Flask ni de app.py."""
    try:
        resultados = buscar_en_documentos(consulta)
        return {
            "cantidad": len(resultados),
            "fragmentos": resultados,
        }
    except Exception as error:
        print("ERROR BUSCANDO EN MANUALES:", error)
        return {"cantidad": 0, "fragmentos": [], "error": "No se pudieron consultar los manuales."}



def _cargar_metadatos():
    """Compatibilidad interna: la persistencia vive en metadata_store.py."""
    return cargar_metadatos()

def _chunks_metadato(contenido, chunk_chars=1400, overlap=220):
    """Divide fichas largas en fragmentos manejables para la recuperación."""
    texto = str(contenido or "").strip()
    if not texto:
        return []
    if len(texto) <= chunk_chars:
        return [texto]
    chunks = []
    inicio = 0
    while inicio < len(texto):
        fin = min(len(texto), inicio + chunk_chars)
        if fin < len(texto):
            corte = max(
                texto.rfind("\n", inicio + 700, fin),
                texto.rfind(". ", inicio + 700, fin),
                texto.rfind("; ", inicio + 700, fin),
            )
            if corte > inicio + 700:
                fin = corte + 1
        fragmento = texto[inicio:fin].strip()
        if fragmento:
            chunks.append(fragmento)
        if fin >= len(texto):
            break
        inicio = max(inicio + 1, fin - overlap)
    return chunks


def _raiz_simple(palabra):
    """Raíz muy simplificada para acercar singular/plural y variaciones
    cercanas (ej. 'grúas'~'grúa', 'remolques'~'remolque') cuando el token
    exacto no matcheó. No es un stemmer real, solo recorta sufijos comunes.
    Esta función faltaba en el archivo original (bug pre-existente) y hacía
    que CUALQUIER búsqueda en metadatos reventara con NameError."""
    p = str(palabra or "")
    for sufijo in ("iciones", "ciones", "mente", "es", "s"):
        if len(p) > len(sufijo) + 3 and p.endswith(sufijo):
            return p[: -len(sufijo)]
    return p


def _puntuar_metadato(consulta, texto):
    consulta_norm = _normalizar_texto(consulta)
    texto_norm = _normalizar_texto(texto)
    if not consulta_norm or not texto_norm:
        return 0
    tokens = [p for p in re.findall(r"[a-z0-9]+", consulta_norm) if len(p) >= 3]
    if not tokens:
        return 0
    tokens = list(_expandir_sinonimos_tokens(tokens))
    puntuacion = 0
    if len(consulta_norm) >= 8 and consulta_norm in texto_norm:
        puntuacion += 30
    palabras_texto = set(re.findall(r"[a-z0-9]+", texto_norm))
    for i in range(len(tokens) - 1):
        if f"{tokens[i]} {tokens[i+1]}" in texto_norm:
            puntuacion += 10
    for token in tokens:
        if token in palabras_texto:
            puntuacion += min(8, 2 + texto_norm.count(token))
        elif _raiz_simple(token) in {_raiz_simple(x) for x in palabras_texto}:
            puntuacion += 3
    return puntuacion


def _companias_mencionadas_en_consulta(consulta):
    """Devuelve compañías visibles mencionadas de forma explícita en la consulta."""
    texto = _normalizar_texto(consulta)
    if not texto:
        return []
    encontradas = []
    for alias, (_codigo, display) in aliases_companias().items():
        alias_norm = _normalizar_texto(alias)
        display_norm = _normalizar_texto(display)
        if not alias_norm and not display_norm:
            continue
        candidatos = {x for x in (alias_norm, display_norm) if x}
        if any(re.search(rf"(?<![a-z0-9]){re.escape(x)}(?![a-z0-9])", texto) for x in candidatos):
            if display and display not in encontradas:
                encontradas.append(display)
    return encontradas


def _ficha_pertenece_a_companias(ficha, companias):
    if not companias:
        return True
    texto = _normalizar_texto(
        f"{ficha.get('titulo', '')}\n{ficha.get('contenido', '')}"
    )
    grupos = _aliases_por_compania_visible()
    for compania in companias:
        aliases = grupos.get(compania, set()) | {_normalizar_texto(compania)}
        if _texto_menciona_compania(texto, aliases):
            return True
    return False


def buscar_en_metadatos(consulta, alcance="puntual"):
    """Busca en las fichas internas con alcance explícito.

    ``puntual`` mantiene un top acotado por relevancia.
    ``exhaustivo`` recorre todas las fichas del universo documental aplicable
    (por compañía cuando puede inferirse) y devuelve evidencia distribuida por
    ficha. La completitud se refiere SIEMPRE a las fichas internas cargadas,
    nunca al universo real de productos de una compañía.
    """
    try:
        fichas = _cargar_metadatos()
    except Exception as error:
        print("ERROR buscar_en_metadatos al cargar:", error)
        return {
            "cantidad": 0,
            "fichas": [],
            "fuente": "Metadatos internos",
            "alcance": alcance,
            "evidencia_estado": "SIN_INFORMACION_SUFICIENTE",
            "error": "No se pudieron cargar los metadatos.",
        }

    alcance = "exhaustivo" if str(alcance or "").lower() == "exhaustivo" else "puntual"
    companias = _companias_mencionadas_en_consulta(consulta)

    # En alcance exhaustivo el universo se restringe a la/s compañía/s explícitas
    # cuando existen; si no, se consideran todas las fichas cargadas.
    universo = [f for f in fichas if _ficha_pertenece_a_companias(f, companias)]
    if not universo and companias:
        # No fingimos exhaustividad sobre un universo vacío por una detección de alias.
        universo = fichas

    resultados = []
    fichas_con_evidencia = set()
    for ficha in universo:
        titulo = str(ficha.get("titulo") or "")
        candidatos_ficha = []
        for fragmento in _chunks_metadato(ficha.get("contenido", "")):
            puntuacion = _puntuar_metadato(consulta, f"{titulo}\n{fragmento}")
            if puntuacion <= 0:
                continue
            candidatos_ficha.append({
                "id": ficha.get("id"),
                "titulo": titulo,
                "contenido": fragmento,
                "actualizado_en": ficha.get("actualizado_en"),
                "puntuacion": puntuacion,
            })
        candidatos_ficha.sort(key=lambda x: x["puntuacion"], reverse=True)
        if candidatos_ficha:
            fichas_con_evidencia.add(ficha.get("id"))
            # Punto: después ordenamos globalmente. Exhaustivo: preservamos
            # representación de cada ficha relevante sin volcar documentos enteros.
            limite_por_ficha = 3 if alcance == "exhaustivo" else len(candidatos_ficha)
            resultados.extend(candidatos_ficha[:limite_por_ficha])

    resultados.sort(key=lambda x: x["puntuacion"], reverse=True)

    salida = []
    vistos = set()
    max_fragmentos = 36 if alcance == "exhaustivo" else 12
    max_chars = 30000 if alcance == "exhaustivo" else 16000
    usados = 0
    for resultado in resultados:
        clave = (resultado["id"], resultado["contenido"])
        if clave in vistos:
            continue
        contenido = str(resultado.get("contenido") or "")
        costo = len(contenido) + len(str(resultado.get("titulo") or "")) + 40
        if salida and usados + costo > max_chars:
            break
        vistos.add(clave)
        salida.append(resultado)
        usados += costo
        if len(salida) >= max_fragmentos:
            break

    if not salida:
        evidencia_estado = "SIN_INFORMACION_SUFICIENTE"
    elif alcance == "exhaustivo" and len(fichas_con_evidencia) < len(universo):
        evidencia_estado = "PARCIAL"
    else:
        evidencia_estado = "CONFIRMADO"

    completitud = {
        "solicitada": alcance == "exhaustivo",
        "universo_fichas_cargadas": len(universo),
        "fichas_revisadas": len(universo),
        "fichas_con_evidencia": len(fichas_con_evidencia),
        "estado": (
            "COMPLETA_SOBRE_FICHAS_CARGADAS"
            if alcance == "exhaustivo"
            else "NO_APLICA"
        ),
        "advertencia": (
            "La completitud describe las fichas internas actualmente cargadas; "
            "no demuestra por sí sola que el catálogo real de la compañía esté completo."
            if alcance == "exhaustivo"
            else ""
        ),
    }

    print(
        f"RETRIEVAL METADATOS: consulta={consulta!r} alcance={alcance} "
        f"fichas_cargadas={len(fichas)} universo={len(universo)} "
        f"fichas_evidencia={len(fichas_con_evidencia)} fragmentos={len(salida)}"
    )
    return {
        "cantidad": len(salida),
        "fichas": salida,
        "fuente": "Metadatos internos",
        "alcance": alcance,
        "companias_detectadas": companias,
        "evidencia_estado": evidencia_estado,
        "completitud": completitud,
    }


# V16: el universo de comparación se deriva de la misma fuente de verdad de
# compañías/alias que usa el resto de OficinaIA. Así una compañía nueva no
# queda afuera del comparador por olvidar agregarla a una segunda lista fija.
def _companias_comparables():
    vistas = []
    for _alias, (_codigo, display) in aliases_companias().items():
        if display and display not in vistas:
            vistas.append(display)
    return tuple(vistas)


def _aliases_por_compania_visible():
    grupos = {nombre: set() for nombre in _companias_comparables()}
    for alias, (_codigo, display) in aliases_companias().items():
        if display in grupos:
            grupos[display].add(_normalizar_texto(alias))
            grupos[display].add(_normalizar_texto(display))
    return grupos


def _texto_menciona_compania(texto_norm, aliases):
    for alias in aliases:
        alias = str(alias or "").strip()
        if not alias:
            continue
        if re.search(rf"(?<![a-z0-9]){re.escape(alias)}(?![a-z0-9])", texto_norm):
            return True
    return False


def _es_consulta_comparativa_companias(pregunta):
    """Detecta consultas de colocación/comparación que deben revisar varias compañías.

    Es deliberadamente acotado: no intenta clasificar todo el chat, sólo evita que una
    pregunta como '¿en qué compañía puedo emitir un auto 56?' termine buscando una
    única compañía o una única ficha.
    """
    t = _normalizar_texto(pregunta)
    if not t:
        return False
    patrones = (
        r"\ben que compania\b",
        r"\ben cuales companias\b",
        r"\bque compania (?:toma|acepta|asegura|emite|cotiza)\b",
        r"\bque companias (?:toman|aceptan|aseguran|emiten|cotizan)\b",
        r"\bdonde (?:puedo )?(?:emitir|asegurar|cotizar|colocar)\b",
        r"\bquien (?:toma|acepta|asegura|emite|cotiza)\b",
        r"\bquienes (?:toman|aceptan|aseguran|emiten|cotizan)\b",
        r"\bcompar(?:a|ame|ar)\b.*\bcompan",
        r"\bcual compania me sirve\b",
        r"\bque compania me sirve\b",
    )
    return any(re.search(p, t) for p in patrones)


def _consulta_comparativa_enriquecida(consulta):
    """Agrega vocabulario de aceptación sin cambiar el riesgo consultado.

    Las fichas suelen hablar de 'antigüedad', 'sin excepción de año', 'admisión' o
    'vehículos aceptados', mientras el usuario dice simplemente 'auto 56'.
    """
    texto = str(consulta or "").strip()
    norm = _normalizar_texto(texto)
    extras = ["acepta", "toma", "emision", "asegurable", "condiciones"]
    if any(x in norm for x in ("auto", "automovil", "vehiculo", "moto", "pickup", "pick up")):
        extras.extend(["ano", "modelo", "antiguedad", "sin excepcion de ano"])
    return f"{texto} {' '.join(extras)}".strip()


def _consulta_comparativa_con_historial(pregunta, historial=None):
    """Completa follow-ups breves usando el último riesgo comparativo del historial.

    Ej.: después de "tengo un auto 56, ¿dónde lo aseguro?", una pregunta
    "¿otras alternativas?" conserva automáticamente "auto 56" como riesgo.
    """
    pregunta = str(pregunta or "").strip()
    historial = historial or []
    norm = _normalizar_texto(pregunta)

    followups = (
        "otras alternativas", "otra alternativa", "alguna otra", "algunas otras",
        "que otra", "que otras", "otra opcion", "otras opciones", "y alguna mas",
        "alguna mas",
    )
    es_followup = any(frase in norm for frase in followups)
    if not es_followup and norm.startswith("y "):
        alias_norm = {_normalizar_texto(a) for a in aliases_companias().keys()}
        es_followup = any(
            re.search(rf"(?<![a-z0-9]){re.escape(a)}(?![a-z0-9])", norm)
            for a in alias_norm if a
        )
    if not es_followup:
        return pregunta

    # Buscamos hacia atrás el último mensaje del usuario con intención de colocación
    # o con un riesgo vehicular concreto.
    for turno in reversed(historial[-12:]):
        if turno.get("rol") != "user":
            continue
        anterior = str(turno.get("contenido") or "").strip()
        if not anterior:
            continue
        anterior_norm = _normalizar_texto(anterior)
        if (
            _es_consulta_comparativa_companias(anterior)
            or re.search(r"\b(auto|automovil|vehiculo|moto|pickup|pick up|camioneta)\b", anterior_norm)
        ):
            return f"{anterior}\nSEGUIMIENTO DEL USUARIO: {pregunta}"

    return pregunta


def _extraer_anio_vehiculo(consulta):
    """Extrae un año/modelo vehicular razonable para validaciones de antigüedad."""
    texto = _normalizar_texto(consulta)
    hoy = datetime.now().year

    # Primero años explícitos de 4 dígitos.
    candidatos = [int(x) for x in re.findall(r"\b(19\d{2}|20\d{2})\b", texto)]
    candidatos = [x for x in candidatos if 1900 <= x <= hoy]
    if candidatos:
        return candidatos[-1]

    # Luego abreviaturas típicas de oficina: "auto 56", "moto 94", "modelo 05".
    m = re.search(
        r"\b(?:auto|automovil|vehiculo|moto|pickup|pick up|camioneta|modelo)\s+(?:modelo\s+)?['’]?(\d{2})\b",
        texto,
    )
    if not m:
        return None

    corto = int(m.group(1))
    corte = hoy % 100
    anio = 2000 + corto if corto <= corte else 1900 + corto
    return anio if 1900 <= anio <= hoy else None


def _limites_antiguedad_en_texto(texto):
    """Devuelve límites máximos de antigüedad expresados de forma inequívoca."""
    norm = _normalizar_texto(texto)
    patrones = (
        r"(?:hasta|maximo|maxima|antiguedad maxima|limite de antiguedad|no mayor a|no superior a)\s*(?:de\s*)?(\d{1,3})\s*anos",
        r"(\d{1,3})\s*anos\s*(?:de antiguedad\s*)?(?:maximo|maxima|como maximo)",
    )
    valores = []
    for pat in patrones:
        for n in re.findall(pat, norm):
            try:
                valores.append(int(n))
            except Exception:
                pass
    return [n for n in valores if 0 < n < 150]


def _evaluar_evidencia_por_antiguedad(consulta, evidencia):
    """Valida compatibilidad temporal cuando la evidencia contiene un límite claro.

    No intenta decidir reglas comerciales complejas: sólo evita contradicciones
    matemáticas como recomendar un vehículo de 70 años donde el máximo es 30.
    """
    anio = _extraer_anio_vehiculo(consulta)
    if not anio:
        return {
            "anio_vehiculo": None,
            "antiguedad_aprox": None,
            "estado": "SIN_VALIDACION_NUMERICA",
            "detalle": "",
        }

    hoy = datetime.now().year
    antiguedad = hoy - anio
    contenido = "\n".join(str(x.get("contenido") or "") for x in evidencia)
    norm = _normalizar_texto(contenido)

    frases_sin_limite = (
        "sin excepcion de ano", "sin excepcion del ano", "sin limite de antiguedad",
        "sin limite de ano", "cualquier ano", "todos los anos",
    )
    if any(frase in norm for frase in frases_sin_limite):
        return {
            "anio_vehiculo": anio,
            "antiguedad_aprox": antiguedad,
            "estado": "COMPATIBLE_POR_ANTIGUEDAD",
            "detalle": f"El vehículo modelo {anio} tiene aproximadamente {antiguedad} años y la evidencia indica que no hay límite/excepción de año.",
        }

    limites = _limites_antiguedad_en_texto(contenido)
    if limites:
        limite = min(limites)
        if antiguedad > limite:
            return {
                "anio_vehiculo": anio,
                "antiguedad_aprox": antiguedad,
                "estado": "NO_COMPATIBLE_POR_ANTIGUEDAD",
                "limite_detectado": limite,
                "detalle": f"El vehículo modelo {anio} tiene aproximadamente {antiguedad} años, que supera el máximo detectado de {limite} años.",
            }
        return {
            "anio_vehiculo": anio,
            "antiguedad_aprox": antiguedad,
            "estado": "COMPATIBLE_POR_LIMITE_DE_ANTIGUEDAD",
            "limite_detectado": limite,
            "detalle": f"El vehículo modelo {anio} tiene aproximadamente {antiguedad} años y no supera el máximo detectado de {limite} años.",
        }

    return {
        "anio_vehiculo": anio,
        "antiguedad_aprox": antiguedad,
        "estado": "SIN_VALIDACION_NUMERICA",
        "detalle": f"Modelo {anio}: antigüedad aproximada {antiguedad} años. La evidencia recuperada no contiene un límite inequívoco de antigüedad.",
    }


def _clasificar_compatibilidad_documental(consulta, evidencia, validacion):
    """Clasifica sólo lo que la evidencia permite afirmar.

    Ausencia de prohibición nunca equivale a aceptación. Para recomendar una
    compañía hace falta una regla positiva de admisión/aceptación o un límite
    de antigüedad explícito compatible con el riesgo consultado.
    """
    if not evidencia:
        return {
            "estado": "SIN_INFORMACION_SUFICIENTE",
            "detalle": "No hay evidencia positiva suficiente para confirmar aceptación.",
        }

    if (validacion or {}).get("estado") == "NO_COMPATIBLE_POR_ANTIGUEDAD":
        return {
            "estado": "NO_COMPATIBLE_CONFIRMADO",
            "detalle": str((validacion or {}).get("detalle") or "La restricción de antigüedad descarta el riesgo."),
        }

    texto = _normalizar_texto("\n".join(str(x.get("contenido") or "") for x in evidencia))
    negativos = (
        r"\bno (?:se )?(?:acepta|aceptan|toma|toman|asegura|aseguran|admite|admiten|emite|emiten)\b",
        r"\b(?:riesgo|vehiculo|vehiculos) no (?:aceptable|asegurable|admisible)\b",
        r"\b(?:queda|quedan) excluid[oa]s?\b",
    )
    if any(re.search(p, texto) for p in negativos):
        return {
            "estado": "NO_COMPATIBLE_CONFIRMADO",
            "detalle": "La evidencia recuperada contiene una exclusión o rechazo explícito.",
        }

    positivos = (
        r"\bsin (?:excepcion|limite).{0,35}(?:ano|antiguedad)\b",
        r"\bcualquier ano\b",
        r"\b(?:se )?(?:acepta|aceptan|toma|toman|admite|admiten)\b",
        r"\b(?:vehiculos?|unidades?) (?:aceptados?|admitidos?|asegurables?)\b",
        r"\b(?:antiguedad maxima|maximo|maxima|hasta)\s*(?:de\s*)?\d{1,3}\s*anos\b",
    )
    hay_regla_positiva = any(re.search(p, texto) for p in positivos)
    estado_validacion = (validacion or {}).get("estado")
    if hay_regla_positiva and estado_validacion in {
        "COMPATIBLE_POR_ANTIGUEDAD",
        "COMPATIBLE_POR_LIMITE_DE_ANTIGUEDAD",
        "SIN_VALIDACION_NUMERICA",
    }:
        return {
            "estado": "COMPATIBLE_CONFIRMADO",
            "detalle": "Hay evidencia positiva de admisión compatible con la restricción temporal detectada.",
        }

    return {
        "estado": "SIN_INFORMACION_SUFICIENTE",
        "detalle": "La ficha puede ser relevante, pero no confirma de forma positiva que este riesgo sea aceptado.",
    }


def comparar_companias(consulta):
    """Recuperación transversal determinística sobre metadatos internos.

    Devuelve evidencia separada por compañía. Una compañía sin evidencia queda como
    'sin evidencia' y jamás como 'no acepta'. Esto permite que Gemini sintetice una
    comparación fiable sin confundir ausencia documental con rechazo comercial.
    """
    try:
        fichas = _cargar_metadatos()
    except Exception as error:
        print("ERROR comparar_companias al cargar metadatos:", error)
        return {
            "cantidad": 0,
            "companias_con_evidencia": 0,
            "companias": [],
            "fuente": "Metadatos internos",
            "error": "No se pudieron cargar los metadatos.",
        }

    consulta_busqueda = _consulta_comparativa_enriquecida(consulta)
    grupos_alias = _aliases_por_compania_visible()
    salida = []

    for compania in _companias_comparables():
        alias_norm = grupos_alias.get(compania, {_normalizar_texto(compania)})
        candidatos = []
        for ficha in fichas:
            titulo = str(ficha.get("titulo") or "")
            contenido_total = str(ficha.get("contenido") or "")
            texto_ficha_norm = _normalizar_texto(f"{titulo}\n{contenido_total}")

            # La ficha debe pertenecer razonablemente a la compañía actual.
            # Esto evita atribuir a ATM una condición que en realidad pertenecía a AGS.
            if not _texto_menciona_compania(texto_ficha_norm, alias_norm):
                continue

            for fragmento in _chunks_metadato(contenido_total):
                score = _puntuar_metadato(
                    f"{compania} {consulta_busqueda}",
                    f"{titulo}\n{fragmento}",
                )
                if score <= 0:
                    continue
                candidatos.append({
                    "id": ficha.get("id"),
                    "titulo": titulo,
                    "contenido": fragmento,
                    "puntuacion": score,
                    "actualizado_en": ficha.get("actualizado_en"),
                })

        candidatos.sort(key=lambda x: x["puntuacion"], reverse=True)
        evidencia = candidatos[:3]
        validacion = _evaluar_evidencia_por_antiguedad(consulta, evidencia) if evidencia else {
            "anio_vehiculo": _extraer_anio_vehiculo(consulta),
            "antiguedad_aprox": (
                datetime.now().year - _extraer_anio_vehiculo(consulta)
                if _extraer_anio_vehiculo(consulta) else None
            ),
            "estado": "SIN_EVIDENCIA",
            "detalle": "No hay evidencia interna suficiente para validar esta compañía.",
        }
        compatibilidad = _clasificar_compatibilidad_documental(consulta, evidencia, validacion)
        salida.append({
            "compania": compania,
            "tiene_evidencia": bool(evidencia),
            "compatibilidad": compatibilidad,
            "validacion": validacion,
            "evidencia": evidencia,
        })

    con_evidencia = sum(1 for item in salida if item["tiene_evidencia"])
    compatibles_confirmadas = sum(
        1 for item in salida
        if (item.get("compatibilidad") or {}).get("estado") == "COMPATIBLE_CONFIRMADO"
    )
    print(
        f"COMPARACION COMPANIAS: consulta={consulta!r} "
        f"companias={len(salida)} con_evidencia={con_evidencia}"
    )
    return {
        "cantidad": con_evidencia,
        "companias_con_evidencia": con_evidencia,
        "companias_compatibles_confirmadas": compatibles_confirmadas,
        "companias_evaluadas": len(salida),
        "companias": salida,
        "fuente": "Metadatos internos por compañía",
        "regla": (
            "Sin evidencia significa información no confirmada; no significa que la compañía rechace el riesgo."
        ),
    }


def _formatear_contexto_comparativo(resultado):
    if not isinstance(resultado, dict):
        return ""
    lineas = [
        "CONSULTA TRANSVERSAL ENTRE COMPAÑÍAS (fuente: metadatos internos):",
        "Regla: 'sin evidencia' NO significa 'no acepta'; sólo significa que no está confirmado con las fichas disponibles.",
    ]
    sin_evidencia = []
    for item in resultado.get("companias", []):
        compania = item.get("compania", "")
        evidencia = item.get("evidencia") or []
        if not evidencia:
            sin_evidencia.append(compania)
            continue
        lineas.append(f"\n{compania}:")
        compatibilidad = item.get("compatibilidad") or {}
        if compatibilidad.get("estado"):
            lineas.append(
                f"- COMPATIBILIDAD DOCUMENTAL: {compatibilidad.get('estado')}"
                + (f" — {compatibilidad.get('detalle')}" if compatibilidad.get("detalle") else "")
            )
        validacion = item.get("validacion") or {}
        if validacion.get("estado"):
            detalle_validacion = str(validacion.get("detalle") or "").strip()
            lineas.append(
                f"- VALIDACION: {validacion.get('estado')}"
                + (f" — {detalle_validacion}" if detalle_validacion else "")
            )
        for ev in evidencia[:2]:
            titulo = str(ev.get("titulo") or "").strip()
            contenido = str(ev.get("contenido") or "").strip()
            if len(contenido) > 900:
                contenido = contenido[:900].rsplit(" ", 1)[0] + "…"
            lineas.append(f"- {titulo}: {contenido}" if titulo else f"- {contenido}")
    if sin_evidencia:
        lineas.append("\nSin evidencia suficiente en metadatos para: " + ", ".join(sin_evidencia) + ".")
    return "\n".join(lineas)


def _buscar_vehiculos_filtrados(datos, compania=None, tipo=None, cliente=None):
    filas = list(datos)
    if compania:
        filas = _filtrar_filas(filas, compania=compania)
    if cliente:
        objetivo = _normalizar_texto(cliente)
        filas = [f for f in filas if objetivo in _normalizar_texto(_valor_campo(f, "CLIENTE") or _valor_campo(f, "ASEGURADO") or "")]
    campo_tipo = _campo_por_alias(filas[0], ("TIPO_VEHICULO", "TIPO DE VEHICULO", "TIPO DE VEHÍCULO", "VEHICULO", "VEHÍCULO", "VH")) if filas else None
    if tipo and campo_tipo:
        objetivos = _expandir_sinonimos_tokens([_normalizar_texto(tipo)])
        filas = [
            f for f in filas
            if any(o in _normalizar_texto(f.get(campo_tipo, "")) for o in objetivos)
        ]
    resultado = []
    vistos = set()
    for f in filas:
        veh = str(_valor_campo(f, "VEHICULO") or _valor_campo(f, "VH") or "").strip()
        pat = str(_valor_campo(f, "PATENTE") or "").strip()
        if not veh and not pat:
            continue
        clave = _normalizar_texto(pat or veh)
        if clave in vistos:
            continue
        vistos.add(clave)
        resultado.append(f)
    return resultado


def buscar_vehiculos(compania=None, tipo=None, cliente=None):
    datos, fuente = _dataset_estructurado()
    filas = _buscar_vehiculos_filtrados(datos, compania=compania, tipo=tipo, cliente=cliente)
    return {"fuente": fuente, "cantidad": len(filas), "vehiculos": filas}


def buscar_en_internet(consulta):
    """Ejecuta una búsqueda web mediante la capacidad de Google Search de Gemini."""
    cliente = obtener_cliente_gemini()
    if cliente is None:
        return {"resultado": "Internet no disponible: falta GEMINI_API_KEY."}
    try:
        config = types.GenerateContentConfig(
            temperature=0.1,
            max_output_tokens=2048,
            tools=[types.Tool(google_search=types.GoogleSearch())],
        )
        respuesta, _modelo = generate_with_fallback(
            client=cliente,
            models=DEFAULT_MODELS[:1],
            contents=consulta,
            config=config,
            log_prefix="BUSQUEDA INTERNET",
        )
        return {"resultado": getattr(respuesta, "text", "") or "No encontré resultados públicos suficientes."}
    except Exception as error:
        print("ERROR BUSQUEDA INTERNET:", error)
        return {"resultado": "No se pudo completar la búsqueda en Internet."}


_TOOL_HANDLERS = {
    "consultar_excel": consultar_excel,
    "contar_registros": contar_registros,
    "analizar_excel": analizar_excel,
    "buscar_en_manuales": buscar_en_manuales,
    "buscar_en_metadatos": buscar_en_metadatos,
    "comparar_companias": comparar_companias,
    "proponer_registro_excel": proponer_registro_excel,
    "guardar_metadato_relevante": guardar_metadato_relevante,
    "buscar_vehiculos": buscar_vehiculos,
    "buscar_en_internet": buscar_en_internet,
}


def _ejecutar_tool(nombre, argumentos, cache=None):
    """Ejecuta una herramienta una sola vez por request cuando es cacheable.

    V16 elimina el concepto de "reintento obligatorio" dirigido por el modelo.
    Una búsqueda vacía se informa como tal y el turno puede terminar sin abrir
    cadenas de llamadas cada vez más caras. El cache vive sólo dentro de
    consultar_gemini(), por lo que nunca contamina requests posteriores.
    """
    handler = _TOOL_HANDLERS.get(nombre)
    if not handler:
        return {"error": f"Herramienta desconocida: {nombre}"}

    herramientas_busqueda = {
        "buscar_en_manuales",
        "buscar_en_metadatos",
        "comparar_companias",
        "consultar_excel",
        "analizar_excel",
        "buscar_vehiculos",
        "buscar_en_internet",
    }
    herramientas_cacheables = {
        "buscar_en_manuales",
        "buscar_en_metadatos",
        "comparar_companias",
        "consultar_excel",
        "analizar_excel",
        "buscar_vehiculos",
    }

    clave_cache = None
    if cache is not None and nombre in herramientas_cacheables:
        try:
            clave_cache = (
                nombre,
                json.dumps(argumentos or {}, ensure_ascii=False, sort_keys=True, default=str),
            )
            if clave_cache in cache:
                return cache[clave_cache]
        except Exception:
            clave_cache = None

    try:
        resultado = handler(**argumentos)
        if isinstance(resultado, dict) and nombre in herramientas_busqueda:
            cantidad = resultado.get("cantidad")
            if cantidad == 0 or resultado.get("error"):
                resultado = dict(resultado)
                resultado["busqueda_vacia"] = True
        if cache is not None and clave_cache is not None:
            cache[clave_cache] = resultado
        return resultado
    except Exception as error:
        print(f"ERROR TOOL {nombre}:", error)
        resultado = {
            "error": f"No se pudo ejecutar {nombre}.",
            "cantidad": 0,
            "busqueda_vacia": True,
        }
        if cache is not None and clave_cache is not None:
            cache[clave_cache] = resultado
        return resultado



def _contenido_respuesta(respuesta):
    texto = getattr(respuesta, "text", None)
    if texto:
        return texto.strip()
    return ""


def _partes_function_calls(respuesta):
    calls = []
    candidatos = getattr(respuesta, "function_calls", None)
    if candidatos:
        for call in candidatos:
            calls.append(call)
        return calls
    for candidato in getattr(respuesta, "candidates", []) or []:
        contenido = getattr(candidato, "content", None)
        for parte in getattr(contenido, "parts", []) or []:
            call = getattr(parte, "function_call", None)
            if call:
                calls.append(call)
    return calls


def _compactar_historial_para_modelo(historial, max_chars=8000, max_por_turno=1800):
    """Convierte el historial visible en memoria conversacional liviana.

    La UI y la base conservan los mensajes completos. Sólo el contexto enviado
    a Gemini se recorta por presupuesto real de caracteres, para que un análisis
    de PDF, un tabulado o una respuesta extensa no siga pesando en todos los
    turnos posteriores. El historial sirve para entender continuaciones; nunca
    representa una operación todavía activa.
    """
    seleccionados = []
    usados = 0
    for turno in reversed((historial or [])[-16:]):
        if not isinstance(turno, dict):
            continue
        rol = turno.get("rol")
        if rol not in {"user", "assistant"}:
            continue
        contenido = str(turno.get("contenido", "") or "").strip()
        if not contenido:
            continue

        # Los bloques tabulados pueden ser enormes y no aportan contexto
        # conversacional en el turno siguiente.
        if contenido.count("\t") >= 8:
            lineas = [ln for ln in contenido.splitlines() if "\t" not in ln]
            contenido = "\n".join(lineas).strip() or "[Resultado tabulado omitido del contexto]"

        if len(contenido) > max_por_turno:
            cabeza = max_por_turno - 280
            contenido = contenido[:cabeza].rstrip() + " … [resumen recortado] … " + contenido[-240:].lstrip()

        linea = f"{'USUARIO' if rol == 'user' else 'ASISTENTE'}: {contenido}"
        costo = len(linea) + 1
        if seleccionados and usados + costo > max_chars:
            break
        if costo > max_chars:
            linea = linea[:max_chars]
            costo = len(linea)
        seleccionados.append(linea)
        usados += costo

    seleccionados.reverse()
    return "\n".join(seleccionados) or "Sin historial relevante."



def _es_followup_breve(pregunta):
    t = _normalizar_texto(pregunta)
    if not t or len(t.split()) > 10:
        return False
    patrones = (
        r"^y\b", r"^y (?:en|para|con)\b", r"^que (?:hay|pasa) con\b",
        r"^y (?:atm|federacion|fed pat|agrosalta|mercantil|rivadavia|san cristobal)\b",
        r"^(?:otra|otras|alguna otra|alguna mas|lo mismo)\b",
    )
    return any(re.search(p, t) for p in patrones)


def _resolver_referente_compania(pregunta, historial=None):
    """Resuelve referencias breves como "esa compañía" sin heredar ejecución.

    Devuelve sólo un nombre/código de compañía; nunca resultados ni herramientas
    del turno anterior. Esto permite cambiar de metadata a Excel conservando el
    referente lingüístico.
    """
    texto = str(pregunta or "").strip()
    n = _normalizar_texto(texto)
    referencias = (
        "esa compania", "esa aseguradora", "de esa compania",
        "esa misma compania", "de esa", "la misma compania",
    )
    if not any(r in n for r in referencias):
        return None

    aliases = aliases_companias()
    # Primero mensajes del usuario, porque suelen fijar el referente más limpio
    # (ej.: "¿Y Federación?"). Luego asistente como fallback.
    turnos = list((historial or [])[-10:])
    for rol in ("user", "assistant"):
        for turno in reversed(turnos):
            if turno.get("rol") != rol:
                continue
            contenido = str(turno.get("contenido") or "")
            contenido_n = _normalizar_texto(contenido)
            mejores = []
            for alias, info in aliases.items():
                alias_n = _normalizar_texto(alias)
                if not alias_n:
                    continue
                m = re.search(rf"(?<![a-z0-9]){re.escape(alias_n)}(?![a-z0-9])", contenido_n)
                if m:
                    codigo, display = info
                    mejores.append((m.start(), len(alias_n), display or codigo))
            if mejores:
                # La última mención específica dentro del turno gana.
                mejores.sort(key=lambda x: (x[0], x[1]))
                return mejores[-1][2]
    return None


def _formatear_contexto_estructurado(resultado):
    if not isinstance(resultado, dict):
        return ""
    if not resultado.get("ok"):
        return "ANÁLISIS ESTRUCTURADO: no se pudo resolver de forma determinística."
    return (
        "===== RESULTADO ESTRUCTURADO EXACTO =====\n"
        + json.dumps(resultado, ensure_ascii=False, indent=2, default=str)
        + "\n===== FIN RESULTADO ESTRUCTURADO ====="
    )


def _construir_plan_ejecucion(pregunta, historial=None):
    """Plan único del turno para fuentes precargadas.

    No ejecuta nada. Sólo decide intención, alcance y consulta documental mínima.
    El historial puede completar follow-ups inequívocos, nunca reactivar ejecuciones.
    """
    historial = historial or []
    referente_compania = _resolver_referente_compania(pregunta, historial)
    pregunta_resuelta = str(pregunta or "").strip()
    if referente_compania:
        pregunta_resuelta = f"{pregunta_resuelta}\n[REFERENTE RESUELTO: {referente_compania}]"

    base = construir_plan_base(pregunta_resuelta).to_dict()
    base["consulta_fuente"] = pregunta_resuelta
    base["contexto_heredado"] = bool(referente_compania)
    if referente_compania:
        base["referente_compania"] = referente_compania

    # La lógica comparativa histórica existente es más rica que el router base.
    pregunta_comparativa = _consulta_comparativa_con_historial(pregunta_resuelta, historial)
    if (
        _es_consulta_comparativa_companias(pregunta_comparativa)
        or pregunta_comparativa != pregunta_resuelta
    ):
        base.update({
            "intencion": "comparacion_companias",
            "alcance": "comparativo",
            "fuentes": ["comparar_companias"],
            "requiere_completitud": True,
            "requiere_metadatos": False,
            "motivo": "consulta transversal de colocación/comparación",
            "consulta_fuente": pregunta_comparativa,
            "contexto_heredado": pregunta_comparativa != str(pregunta or "").strip(),
        })
        return base

    # Follow-up documental corto: heredamos sólo la consulta mínima de la última
    # pregunta documental. No heredamos PDF, /flota, resultados ni herramientas.
    if base.get("intencion") == "general" and _es_followup_breve(pregunta):
        for turno in reversed(historial[-8:]):
            if turno.get("rol") != "user":
                continue
            anterior = str(turno.get("contenido") or "").strip()
            if not anterior or anterior == str(pregunta or "").strip():
                continue
            anterior_plan = construir_plan_base(anterior).to_dict()
            if anterior_plan.get("intencion") == "consulta_documental":
                consulta = f"{anterior}\nSEGUIMIENTO DEL USUARIO: {pregunta}"
                alcance_actual = construir_plan_base(pregunta).to_dict().get("alcance")
                alcance = "exhaustivo" if alcance_actual == "exhaustivo" else anterior_plan.get("alcance", "puntual")
                base.update({
                    "intencion": "consulta_documental",
                    "alcance": alcance,
                    "fuentes": ["buscar_en_metadatos"],
                    "requiere_completitud": alcance == "exhaustivo",
                    "requiere_metadatos": True,
                    "motivo": "seguimiento inequívoco de consulta documental anterior",
                    "consulta_fuente": consulta,
                    "contexto_heredado": True,
                })
                break
    return base


def _formatear_contexto_metadatos(resultado):
    if not isinstance(resultado, dict):
        return ""
    fichas = resultado.get("fichas") or []
    alcance = resultado.get("alcance") or "puntual"
    evidencia = resultado.get("evidencia_estado") or "SIN_INFORMACION_SUFICIENTE"
    completitud = resultado.get("completitud") or {}

    partes = [
        "===== EVIDENCIA DOCUMENTAL INTERNA =====",
        f"ALCANCE DE RECUPERACIÓN: {alcance}",
        f"ESTADO DE EVIDENCIA: {evidencia}",
    ]
    if alcance == "exhaustivo":
        partes.extend([
            "COMPLETITUD SOBRE FICHAS CARGADAS:",
            f"- fichas del universo revisado: {completitud.get('universo_fichas_cargadas', 0)}",
            f"- fichas con evidencia temática: {completitud.get('fichas_con_evidencia', 0)}",
            f"- estado del barrido: {completitud.get('estado', 'DESCONOCIDO')}",
            "IMPORTANTE: un barrido completo de las fichas cargadas NO demuestra que el catálogo real de la compañía esté completo.",
        ])
    if not fichas:
        partes.extend([
            "No se encontró evidencia documental suficiente en las fichas internas cargadas.",
            "===== FIN EVIDENCIA DOCUMENTAL INTERNA =====",
        ])
        return "\n".join(partes)

    partes.append("")
    for ficha in fichas:
        titulo = str(ficha.get("titulo") or "Metadato").strip()
        contenido = str(ficha.get("contenido") or "").strip()
        if not contenido:
            continue
        partes.append(f"[Ficha: {titulo}]")
        partes.append(contenido)
        partes.append("")
    partes.append("===== FIN EVIDENCIA DOCUMENTAL INTERNA =====")
    return "\n".join(partes)


def _plan_para_prompt(plan):
    fuentes = ", ".join(plan.get("fuentes") or []) or "ninguna precargada"
    return (
        f"INTENCIÓN: {plan.get('intencion', 'general')}\n"
        f"ALCANCE: {plan.get('alcance', 'puntual')}\n"
        f"FUENTES PLANIFICADAS: {fuentes}\n"
        f"REQUIERE COMPLETITUD: {'sí' if plan.get('requiere_completitud') else 'no'}\n"
        f"CONTEXTO HEREDADO: {'sí' if plan.get('contexto_heredado') else 'no'}\n"
        f"REFERENTE DE COMPAÑÍA: {plan.get('referente_compania') or 'ninguno'}\n"
        f"MOTIVO: {plan.get('motivo', '')}"
    )

def _tools_para_plan(plan):
    """Evita ofrecer de nuevo a Gemini una fuente que el plan ya ejecutó.

    Sofia conserva las demás herramientas para operaciones determinísticas o
    ampliaciones excepcionales. Esto elimina el doble owner precontexto/tool.
    """
    ejecutadas = set(plan.get("fuentes") or [])
    excluir = ejecutadas & {"buscar_en_metadatos", "comparar_companias", "analizar_excel"}
    if not excluir:
        return TOOL_DEFINITIONS

    salida = []
    for tool in TOOL_DEFINITIONS:
        declaraciones = list(getattr(tool, "function_declarations", None) or [])
        filtradas = [d for d in declaraciones if getattr(d, "name", "") not in excluir]
        if filtradas:
            salida.append(types.Tool(function_declarations=filtradas))
    return salida or TOOL_DEFINITIONS


def consultar_gemini(pregunta, contexto="", historial=None):
    cliente = obtener_cliente_gemini()
    if cliente is None:
        return "La IA todavía no está configurada. Falta GEMINI_API_KEY."

    historial = historial or []
    historial_texto = _compactar_historial_para_modelo(historial)
    # Cache estrictamente efímero: nace y muere dentro de este turno.
    tool_cache = {}

    # V20 Etapa 2: un único plan decide qué contexto precargar antes de Gemini.
    # app.py ya no hace pre-routing documental por su cuenta.
    plan = _construir_plan_ejecucion(pregunta, historial)
    print(
        "EXECUTION PLAN:",
        {k: plan.get(k) for k in (
            "intencion", "alcance", "fuentes", "requiere_completitud",
            "contexto_heredado", "motivo"
        )}
    )

    contexto_comparativo = ""
    contexto_documental_plan = ""
    contexto_estructurado = ""
    consulta_fuente = plan.get("consulta_fuente") or str(pregunta or "").strip()

    if "comparar_companias" in (plan.get("fuentes") or []):
        try:
            resultado_comparativo = _ejecutar_tool(
                "comparar_companias", {"consulta": consulta_fuente}, cache=tool_cache
            )
            contexto_comparativo = _formatear_contexto_comparativo(resultado_comparativo)
        except Exception as error:
            print("ERROR PRECONTEXTO COMPARATIVO:", error)

    if "analizar_excel" in (plan.get("fuentes") or []):
        try:
            resultado_estructurado = _ejecutar_tool(
                "analizar_excel", {"consulta": consulta_fuente}, cache=tool_cache
            )
            contexto_estructurado = _formatear_contexto_estructurado(resultado_estructurado)
        except Exception as error:
            print("ERROR PRECONTEXTO ANALITICO EXCEL:", error)

    if "contar_registros" in (plan.get("fuentes") or []) and plan.get("referente_compania"):
        try:
            qn = _normalizar_texto(pregunta)
            tipo_conteo = "unicos" if any(x in qn for x in ("asegurado", "persona", "cliente")) else "filas"
            resultado_conteo = _ejecutar_tool(
                "contar_registros",
                {"compania": plan.get("referente_compania"), "tipo_conteo": tipo_conteo},
                cache=tool_cache,
            )
            contexto_estructurado = _formatear_contexto_estructurado({
                "ok": True,
                "operacion": "conteo_con_referente",
                "referente_compania": plan.get("referente_compania"),
                **(resultado_conteo if isinstance(resultado_conteo, dict) else {"resultado": resultado_conteo}),
            })
        except Exception as error:
            print("ERROR PRECONTEXTO CONTEO REFERENTE:", error)

    if "buscar_en_metadatos" in (plan.get("fuentes") or []):
        try:
            alcance_busqueda = (
                "exhaustivo" if plan.get("alcance") == "exhaustivo" else "puntual"
            )
            resultado_metadatos = _ejecutar_tool(
                "buscar_en_metadatos",
                {"consulta": consulta_fuente, "alcance": alcance_busqueda},
                cache=tool_cache,
            )
            contexto_documental_plan = _formatear_contexto_metadatos(resultado_metadatos)
        except Exception as error:
            print("ERROR PRECONTEXTO METADATOS:", error)

    contexto_documental = "\n\n".join(
        parte for parte in (str(contexto or "").strip(), contexto_documental_plan.strip())
        if parte
    )

    fecha_hoy = datetime.now().strftime("%d/%m/%Y")
    prompt = build_sofia_prompt(
        fecha_hoy=fecha_hoy,
        plan_texto=_plan_para_prompt(plan),
        historial_texto=historial_texto,
        contexto_comparativo=contexto_comparativo,
        contexto_documental=contexto_documental,
        contexto_estructurado=contexto_estructurado,
        pregunta=pregunta,
    )

    contents = [prompt]
    propuesta_excel = None
    propuesta_metadato = None

    # V16: las búsquedas vacías ya no abren un "reintento obligatorio".
    # El modelo recibe el resultado vacío/error y puede responder con claridad.
    # Esto evita que las consultas difíciles sean justamente las que más
    # llamadas encadenen y más se acerquen al timeout del único worker.
    total_tool_calls = 0
    MAX_TOOL_CALLS = 8

    # Antes eran 6 vueltas x hasta 3 modelos cada una (hasta 18 llamadas a
    # Gemini encadenadas en un mismo request). Con timeout de gunicorn en
    # 180s, eso terminaba en 502 (timeout) o 500. Con los sinónimos de dominio
    # ya aplicados en la búsqueda, la primera consulta casi siempre encuentra
    # el dato; 3 vueltas alcanzan de sobra para el flujo normal + 1 reintento.
    LIMITE_VUELTAS = 3
    for _ in range(LIMITE_VUELTAS):
        ultimo_error = None
        respuesta = None

        try:
            config = types.GenerateContentConfig(
                temperature=0.05,
                max_output_tokens=4096,
                tools=_tools_para_plan(plan),
                # OficinaIA administra manualmente el ciclo Gemini -> tool -> Gemini.
                # Desactivar AFC evita tener dos orquestadores superpuestos.
                automatic_function_calling=types.AutomaticFunctionCallingConfig(
                    disable=True
                ),
            )
            respuesta, _modelo_usado = generate_with_fallback(
                client=cliente,
                models=DEFAULT_MODELS,
                contents=contents,
                config=config,
                log_prefix="GEMINI",
            )
        except Exception as error:
            ultimo_error = error
            respuesta = None

        if respuesta is None:
            print("GEMINI TODOS LOS MODELOS FALLARON:", ultimo_error)
            if propuesta_excel or propuesta_metadato:
                return (
                    "Gemini no está disponible en este momento. La propuesta quedó "
                    "pendiente de confirmación.",
                    propuesta_excel,
                    propuesta_metadato,
                )
            return "Gemini no está disponible en este momento. Intentá nuevamente en unos segundos."

        calls = _partes_function_calls(respuesta)
        if not calls:
            texto = _contenido_respuesta(respuesta) or "No pude generar una respuesta con la información disponible."

            if propuesta_excel or propuesta_metadato:
                return texto, propuesta_excel, propuesta_metadato
            return texto

        contents.append(
            respuesta.candidates[0].content
            if getattr(respuesta, "candidates", None)
            else respuesta
        )

        for call in calls:
            nombre = getattr(call, "name", "")
            argumentos = dict(getattr(call, "args", {}) or {})
            print("GEMINI TOOL CALL:", nombre, argumentos)

            total_tool_calls += 1
            if total_tool_calls > MAX_TOOL_CALLS:
                resultado = {
                    "error": "Se alcanzó el límite de herramientas para este turno.",
                    "cantidad": 0,
                }
            else:
                resultado = _ejecutar_tool(nombre, argumentos, cache=tool_cache)

            if nombre == "proponer_registro_excel":
                propuesta_excel = (
                    resultado.get("propuesta")
                    if isinstance(resultado, dict)
                    else None
                )

            if nombre == "guardar_metadato_relevante":
                propuesta_metadato = (
                    resultado.get("propuesta")
                    if isinstance(resultado, dict)
                    else None
                )

            contents.append(types.Part.from_function_response(
                name=nombre,
                response={"resultado": resultado},
            ))

    if propuesta_excel or propuesta_metadato:
        return (
            "No pude completar la consulta después de consultar las fuentes disponibles.",
            propuesta_excel,
            propuesta_metadato,
        )
    return "No pude completar la consulta después de consultar las fuentes disponibles."
